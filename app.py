from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from authlib.integrations.flask_client import OAuth
from datetime import datetime, date
from decimal import Decimal
import os
import secrets
from dotenv import load_dotenv
import os

from sqlalchemy import func, desc, or_, and_
from models import (
    db, User, UserRole, AcademicYear, Class, Stream, Vehicle, Student, StudentType,
    FeeItem, FeeRate, FeeScope, StudentFeeAssignment, FeeAssessment, Payment,
    PaymentMode, PaymentAllocation, StudentPromotion, PromotionStatus,
    ExpenseCategory, Expense, generate_fee_assessments, get_student_balance_summary
)

# Dynamically find and load the .env file from current project directory
basedir = os.path.abspath(os.path.dirname(__file__))
load_dotenv(os.path.join(basedir, '.env'))

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(16))
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///school_fees.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Google OAuth configuration
app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET')


# Initialize extensions
db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
oauth = OAuth(app)

# Configure Google OAuth
google = oauth.register(
    name='google',
    client_id=app.config['GOOGLE_CLIENT_ID'],
    client_secret=app.config['GOOGLE_CLIENT_SECRET'],
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile'
    }
)



@app.template_global()
def get_date():
    """Make date class available in templates"""
    return date

@app.template_global()
def get_datetime():
    """Make datetime class available in templates"""
    return datetime

@app.template_global()
def get_decimal():
    """Make Decimal class available in templates"""
    return Decimal

# Alternative approach - inject into ALL template contexts at once
@app.context_processor
def inject_common_vars():
    """Inject common variables into all template contexts"""
    return {
        'date': date,
        'datetime': datetime,
        'Decimal': Decimal,
        'today': date.today(),  # Current date
        'now': datetime.now()   # Current datetime
    }

# You can also add your enums if templates need them
from models import StudentType, UserRole, PaymentMode, FeeScope, PromotionStatus

@app.context_processor
def inject_enums():
    """Inject enum classes into all template contexts"""
    return {
        'StudentType': StudentType,
        'UserRole': UserRole,
        'PaymentMode': PaymentMode,
        'FeeScope': FeeScope,
        'PromotionStatus': PromotionStatus
    }

# If you prefer the template_global approach for enums:
@app.template_global()
def get_student_type():
    return StudentType

@app.template_global()
def get_user_role():
    return UserRole

@app.template_global()
def get_payment_mode():
    return PaymentMode

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ===========================
#  AUTHENTICATION ROUTES
# ===========================
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('auth/login.html')


@app.route('/login')
def login():
    redirect_uri = url_for('auth_callback', _external=True)
    return google.authorize_redirect(redirect_uri)


@app.route('/auth/callback')
def auth_callback():
    token = google.authorize_access_token()
    user_info = token.get('userinfo')

    if user_info:
        user = User.query.filter_by(email=user_info['email']).first()

        if not user:
            user = User(
                google_id=user_info['sub'],
                email=user_info['email'],
                name=user_info['name'],
                profile_pic=user_info.get('picture'),
                role=UserRole.ACCOUNTANT  # Default role
            )
            db.session.add(user)
            db.session.commit()

        user.last_login = datetime.utcnow()
        db.session.commit()

        login_user(user)
        return redirect(url_for('dashboard'))

    flash('Authentication failed', 'error')
    return redirect(url_for('index'))


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))


# ===========================
#  DASHBOARD
# ===========================
@app.route('/dashboard')
@login_required
def dashboard():
    # Get current academic year
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    # Basic statistics
    stats = {
        'total_students': Student.query.filter_by(is_active=True).count(),
        'total_classes': Class.query.count(),
        'total_vehicles': Vehicle.query.filter_by(is_active=True).count(),
        'current_year': current_year.year if current_year else 'Not Set'
    }

    # Recent payments (last 10)
    recent_payments = Payment.query.order_by(desc(Payment.created_at)).limit(10).all()

    # Students with outstanding balances
    students_with_balances = []
    students = Student.query.filter_by(is_active=True).limit(20).all()  # Limit for performance

    for student in students:
        balance = student.get_current_balance()
        if balance > 0:
            students_with_balances.append({
                'student': student,
                'balance': balance
            })

    # Sort by balance descending
    students_with_balances.sort(key=lambda x: x['balance'], reverse=True)
    students_with_balances = students_with_balances[:10]  # Top 10 outstanding

    return render_template('dashboard.html',
                           stats=stats,
                           recent_payments=recent_payments,
                           students_with_balances=students_with_balances)


# ===========================
#  ACADEMIC MANAGEMENT
# ===========================
@app.route('/academic')
@login_required
def academic_management():
    classes = Class.query.all()
    academic_years = AcademicYear.query.order_by(desc(AcademicYear.year)).all()
    return render_template('academic/index.html', classes=classes, academic_years=academic_years)


@app.route('/academic/class/add', methods=['GET', 'POST'])
@login_required
def add_class():
    if request.method == 'POST':
        class_obj = Class(
            name=request.form['name'],
            level=request.form['level'],
            next_class_id=request.form['next_class_id'] if request.form['next_class_id'] else None
        )
        db.session.add(class_obj)
        db.session.commit()
        flash('Class added successfully', 'success')
        return redirect(url_for('academic_management'))

    classes = Class.query.all()
    return render_template('academic/add_class.html', classes=classes)


@app.route('/academic/stream/add/<int:class_id>', methods=['GET', 'POST'])
@login_required
def add_stream(class_id):
    class_obj = Class.query.get_or_404(class_id)

    if request.method == 'POST':
        stream = Stream(
            class_id=class_id,
            name=request.form['name'],
            capacity=int(request.form['capacity']) if request.form['capacity'] else 40
        )
        db.session.add(stream)
        db.session.commit()
        flash('Stream added successfully', 'success')
        return redirect(url_for('academic_management'))

    return render_template('academic/add_stream.html', class_obj=class_obj)


@app.route('/academic/year/add', methods=['GET', 'POST'])
@login_required
def add_academic_year():
    if request.method == 'POST':
        year = AcademicYear(
            year=int(request.form['year']),
            start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date(),
            is_current=bool(request.form.get('is_current'))
        )

        if year.is_current:
            # Unset other current years
            AcademicYear.query.update({'is_current': False})

        db.session.add(year)
        db.session.commit()
        flash('Academic year added successfully', 'success')
        return redirect(url_for('academic_management'))

    return render_template('academic/add_year.html')


# ===========================
#  STUDENT MANAGEMENT
# ===========================
@app.route('/students')
@login_required
def student_list():
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # Search and filter
    search = request.args.get('search', '')
    class_id = request.args.get('class_id', type=int)
    stream_id = request.args.get('stream_id', type=int)

    query = Student.query.filter_by(is_active=True)

    if search:
        query = query.filter(or_(
            Student.admission_no.contains(search),
            Student.first_name.contains(search),
            Student.last_name.contains(search)
        ))

    if class_id:
        query = query.filter_by(class_id=class_id)

    if stream_id:
        query = query.filter_by(stream_id=stream_id)

    students = query.paginate(page=page, per_page=per_page, error_out=False)
    classes = Class.query.all()
    streams = Stream.query.filter_by(class_id=class_id).all() if class_id else []

    return render_template('students/list.html',
                           students=students,
                           classes=classes,
                           streams=streams,
                           search=search,
                           selected_class=class_id,
                           selected_stream=stream_id)


@app.route('/students/add', methods=['GET', 'POST'])
@login_required
def add_student():
    if request.method == 'POST':
        student = Student(
            admission_no=request.form['admission_no'],
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            date_of_birth=datetime.strptime(request.form['date_of_birth'], '%Y-%m-%d').date() if request.form[
                'date_of_birth'] else None,
            class_id=int(request.form['class_id']),
            stream_id=int(request.form['stream_id']) if request.form['stream_id'] else None,
            student_type=StudentType(request.form['student_type']),
            parent_name=request.form['parent_name'],
            parent_phone=request.form['parent_phone'],
            parent_email=request.form['parent_email'],
            vehicle_id=int(request.form['vehicle_id']) if request.form['vehicle_id'] else None,
            transport_distance_km=Decimal(request.form['transport_distance_km']) if request.form[
                'transport_distance_km'] else None
        )

        db.session.add(student)
        db.session.commit()
        flash('Student added successfully', 'success')
        return redirect(url_for('student_list'))

    classes = Class.query.all()
    vehicles = Vehicle.query.filter_by(is_active=True).all()
    return render_template('students/add.html', classes=classes, vehicles=vehicles)


@app.route('/students/<int:student_id>')
@login_required
def student_detail(student_id):
    student = Student.query.get_or_404(student_id)
    balance_summary = get_student_balance_summary(student_id)

    # Get recent payments
    recent_payments = Payment.query.filter_by(student_id=student_id) \
        .order_by(desc(Payment.payment_date)).limit(10).all()

    # Get current term assessments
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    current_assessments = []
    if current_year:
        current_assessments = FeeAssessment.query.filter_by(
            student_id=student_id,
            year=current_year.year
        ).all()

    return render_template('students/detail.html',
                           student=student,
                           balance_summary=balance_summary,
                           recent_payments=recent_payments,
                           current_assessments=current_assessments)


@app.route('/api/streams/<int:class_id>')
@login_required
def get_streams(class_id):
    """API endpoint to get streams for a class"""
    streams = Stream.query.filter_by(class_id=class_id).all()
    return jsonify([{'id': s.id, 'name': s.name} for s in streams])


# ===========================
#  VEHICLE MANAGEMENT
# ===========================
@app.route('/vehicles')
@login_required
def vehicle_list():
    vehicles = Vehicle.query.all()
    return render_template('vehicles/list.html', vehicles=vehicles)


@app.route('/vehicles/add', methods=['GET', 'POST'])
@login_required
def add_vehicle():
    if request.method == 'POST':
        vehicle = Vehicle(
            registration_number=request.form['registration_number'],
            make=request.form['make'],
            model=request.form['model'],
            capacity=int(request.form['capacity']) if request.form['capacity'] else None,
            driver_name=request.form['driver_name'],
            driver_phone=request.form['driver_phone'],
            route_description=request.form['route_description']
        )

        db.session.add(vehicle)
        db.session.commit()
        flash('Vehicle added successfully', 'success')
        return redirect(url_for('vehicle_list'))

    return render_template('vehicles/add.html')


@app.route('/vehicles/<int:vehicle_id>')
@login_required
def vehicle_detail(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    students = Student.query.filter_by(vehicle_id=vehicle_id, is_active=True).all()

    # Calculate total revenue for this vehicle
    total_revenue = 0
    for student in students:
        if student.transport_distance_km:
            # Get current transport rate (simplified)
            current_year = AcademicYear.query.filter_by(is_current=True).first()
            if current_year:
                transport_fee = FeeItem.query.filter_by(code='TRANSPORT').first()
                if transport_fee:
                    rate = FeeRate.query.filter_by(
                        fee_item_id=transport_fee.id,
                        year=current_year.year
                    ).first()
                    if rate and rate.rate_per_km:
                        total_revenue += float(student.transport_distance_km * rate.rate_per_km * 3)  # 3 terms

    return render_template('vehicles/detail.html',
                           vehicle=vehicle,
                           students=students,
                           total_revenue=total_revenue)


# ===========================
#  FEE MANAGEMENT
# ===========================
@app.route('/fees')
@login_required
def fee_management():
    # Get filter parameters
    term = request.args.get('term', 1, type=int)
    year = request.args.get('year', type=int)

    current_year = AcademicYear.query.filter_by(is_current=True).first()

    # Default to current year if not specified
    if not year and current_year:
        year = current_year.year
    elif not year:
        year = 2025

    # Get all fee items
    fee_items = FeeItem.query.all()

    # Get current rates for the selected term/year
    current_rates = []
    if term and year:
        current_rates = FeeRate.query.filter_by(
            term=term,
            year=year,
            is_active=True
        ).all()

    # Get classes for the modal
    classes = Class.query.all()

    return render_template('fees/index.html',
                           fee_items=fee_items,
                           current_rates=current_rates,
                           current_year=current_year,
                           classes=classes)

@app.route('/fees/item/add', methods=['GET', 'POST'])
@login_required
def add_fee_item():
    if request.method == 'POST':
        fee_item = FeeItem(
            name=request.form['name'],
            code=request.form['code'].upper(),
            description=request.form['description'],
            scope=FeeScope(request.form['scope']),
            is_per_km=bool(request.form.get('is_per_km'))
        )

        db.session.add(fee_item)
        db.session.commit()
        flash('Fee item added successfully', 'success')
        return redirect(url_for('fee_management'))

    return render_template('fees/add_item.html')


@app.route('/fees/rate/add', methods=['GET', 'POST'])
@login_required
def add_fee_rate():
    if request.method == 'POST':
        rate = FeeRate(
            fee_item_id=int(request.form['fee_item_id']),
            term=int(request.form['term']),
            year=int(request.form['year']),
            class_id=int(request.form['class_id']) if request.form.get('class_id') else None,
            stream_id=int(request.form['stream_id']) if request.form.get('stream_id') else None,  # Use .get() here
            student_type=StudentType(request.form['student_type']) if request.form.get('student_type') else None,
            amount=Decimal(request.form['amount']) if request.form.get('amount') else None,
            rate_per_km=Decimal(request.form['rate_per_km']) if request.form.get('rate_per_km') else None
        )

        db.session.add(rate)
        db.session.commit()
        flash('Fee rate added successfully', 'success')
        return redirect(url_for('fee_management'))

    fee_items = FeeItem.query.filter_by(is_active=True).all()
    classes = Class.query.all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('fees/add_rate.html',
                           fee_items=fee_items,
                           classes=classes,
                           current_year=current_year)

# Replace the existing assess_fees route with this enhanced version

@app.route('/fees/assess', methods=['GET', 'POST'])
@login_required
def assess_fees():
    """Enhanced assess fees with preview functionality"""
    if request.method == 'POST':
        # Check if this is a preview request
        if request.form.get('preview'):
            return handle_fee_assessment_preview()
        else:
            # Actual assessment generation
            term = int(request.form['term'])
            year = int(request.form['year'])
            assessment_scope = request.form.get('assessment_scope', 'all')

            class_id = None
            stream_id = None
            student_id = None

            if assessment_scope == 'class':
                class_id = int(request.form['class_id']) if request.form.get('class_id') else None
            elif assessment_scope == 'stream':
                class_id = int(request.form['class_id']) if request.form.get('class_id') else None
                stream_id = int(request.form['stream_id']) if request.form.get('stream_id') else None
            elif assessment_scope == 'individual':
                student_id = int(request.form['student_id']) if request.form.get('student_id') else None

            # Check if this is a dry run
            dry_run = bool(request.form.get('dry_run'))

            # Check if force regenerate is enabled
            force_regenerate = bool(request.form.get('force_regenerate'))

            if not dry_run:
                assessments_created = generate_fee_assessments(
                    term=term,
                    year=year,
                    class_id=class_id,
                    stream_id=stream_id,
                    student_id=student_id,
                    force_regenerate=force_regenerate
                )

                if force_regenerate:
                    flash(f'Regenerated {assessments_created} fee assessments', 'success')
                else:
                    flash(f'Generated {assessments_created} fee assessments', 'success')
            else:
                flash('Dry run completed - no assessments were created', 'info')

            return redirect(url_for('fee_management'))

    classes = Class.query.all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('fees/assess.html',
                           classes=classes,
                           current_year=current_year)

def handle_fee_assessment_preview():
    """Handle preview request for fee assessments"""
    try:
        term = int(request.form['term'])
        year = int(request.form['year'])
        assessment_scope = request.form.get('assessment_scope', 'all')
        skip_existing = bool(request.form.get('skip_existing'))
        include_transport = bool(request.form.get('include_transport'))

        # Determine target students
        query = Student.query.filter_by(is_active=True)

        if assessment_scope == 'class' and request.form.get('class_id'):
            query = query.filter_by(class_id=int(request.form['class_id']))
        elif assessment_scope == 'stream' and request.form.get('stream_id'):
            query = query.filter_by(stream_id=int(request.form['stream_id']))
        elif assessment_scope == 'individual' and request.form.get('student_id'):
            query = query.filter_by(id=int(request.form['student_id']))

        students = query.all()

        # Simulate assessment generation
        preview_data = {
            'students_count': len(students),
            'assessments_count': 0,
            'total_amount': 'KSh 0.00',
            'skipped_count': 0,
            'fee_breakdown': [],
            'warnings': []
        }

        total_amount = Decimal('0')
        fee_totals = {}

        for student in students:
            # Check existing assessments if skip_existing is enabled
            if skip_existing:
                existing = FeeAssessment.query.filter_by(
                    student_id=student.id,
                    term=term,
                    year=year
                ).first()

                if existing:
                    preview_data['skipped_count'] += 1
                    continue

            # Get applicable fees for this student
            applicable_fees = get_applicable_fees_for_student(student, term, year)

            for fee_item, rate_info in applicable_fees:
                # Skip transport fees if not included
                if not include_transport and fee_item.code == 'TRANSPORT':
                    continue

                amount = calculate_fee_amount(student, fee_item, rate_info)

                if amount > 0:
                    preview_data['assessments_count'] += 1
                    total_amount += amount

                    # Track fee breakdown
                    if fee_item.code not in fee_totals:
                        fee_totals[fee_item.code] = {
                            'name': fee_item.name,
                            'code': fee_item.code,
                            'students_count': 0,
                            'amount': Decimal('0')
                        }

                    fee_totals[fee_item.code]['students_count'] += 1
                    fee_totals[fee_item.code]['amount'] += amount

        # Format fee breakdown
        for code, data in fee_totals.items():
            preview_data['fee_breakdown'].append({
                'code': data['code'],
                'name': data['name'],
                'students_count': data['students_count'],
                'amount': f"KSh {data['amount']:,.2f}"
            })

        preview_data['total_amount'] = f"KSh {total_amount:,.2f}"

        # Add warnings
        if preview_data['assessments_count'] == 0:
            preview_data['warnings'].append('No new assessments will be created with the current criteria.')

        if not include_transport:
            transport_students = sum(1 for s in students if s.vehicle_id)
            if transport_students > 0:
                preview_data['warnings'].append(
                    f'{transport_students} students have transport assignments but transport fees are excluded.')

        return jsonify({'success': True, 'data': preview_data})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Preview failed: {str(e)}'})


# Also add these helper functions if they don't exist in your models.py

def get_applicable_fees_for_student(student, term, year):
    """Get all fee items applicable to a specific student"""
    applicable_fees = []

    # Get all active fee items
    fee_items = FeeItem.query.filter_by(is_active=True).all()

    for fee_item in fee_items:
        rate_info = None

        # Check for individual assignment first
        individual_assignment = StudentFeeAssignment.query.filter_by(
            student_id=student.id,
            fee_item_id=fee_item.id,
            term=term,
            year=year,
            is_active=True
        ).first()

        if individual_assignment:
            if individual_assignment.custom_amount:
                rate_info = {
                    'source': 'individual',
                    'base_rate': individual_assignment.custom_amount,
                    'quantity': 1
                }
            elif fee_item.is_per_km and individual_assignment.custom_rate_per_km:
                distance = individual_assignment.custom_distance or student.transport_distance_km
                if distance:
                    rate_info = {
                        'source': 'individual_per_km',
                        'base_rate': individual_assignment.custom_rate_per_km,
                        'quantity': distance
                    }

        # If no individual assignment found, look for standard rates
        # THIS IS THE KEY FIX - check standard rates regardless of scope
        if not rate_info:
            if fee_item.scope == FeeScope.UNIVERSAL or fee_item.scope == FeeScope.INDIVIDUAL:
                # Universal fees OR individual-scope fees with universal rates
                rate = FeeRate.query.filter_by(
                    fee_item_id=fee_item.id,
                    term=term,
                    year=year,
                    class_id=None,
                    stream_id=None,
                    is_active=True
                ).first()

                if rate:
                    rate_info = get_rate_info(rate, student, fee_item)

                    # Additional check for transport: only apply if student has vehicle
                    if fee_item.code == 'TRANSPORT' and not student.vehicle_id:
                        rate_info = None

            elif fee_item.scope == FeeScope.STREAM_LEVEL and student.stream_id:
                # ... rest of stream logic
                rate = FeeRate.query.filter_by(
                    fee_item_id=fee_item.id,
                    term=term,
                    year=year,
                    stream_id=student.stream_id,
                    student_type=student.student_type,
                    is_active=True
                ).first()

                if not rate:
                    rate = FeeRate.query.filter_by(
                        fee_item_id=fee_item.id,
                        term=term,
                        year=year,
                        stream_id=student.stream_id,
                        student_type=None,
                        is_active=True
                    ).first()

                if rate:
                    rate_info = get_rate_info(rate, student, fee_item)

            elif fee_item.scope == FeeScope.CLASS_LEVEL:
                # ... rest of class logic
                rate = FeeRate.query.filter_by(
                    fee_item_id=fee_item.id,
                    term=term,
                    year=year,
                    class_id=student.class_id,
                    student_type=student.student_type,
                    is_active=True
                ).first()

                if not rate:
                    rate = FeeRate.query.filter_by(
                        fee_item_id=fee_item.id,
                        term=term,
                        year=year,
                        class_id=student.class_id,
                        student_type=None,
                        is_active=True
                    ).first()

                if rate:
                    rate_info = get_rate_info(rate, student, fee_item)

        if rate_info:
            applicable_fees.append((fee_item, rate_info))

    return applicable_fees


def generate_fee_assessments(term, year, class_id=None, stream_id=None, student_id=None, force_regenerate=False):
    """Generate fee assessments for students based on their applicable fees

    Args:
        term: Academic term (1, 2, or 3)
        year: Academic year
        class_id: Optional - limit to specific class
        stream_id: Optional - limit to specific stream
        student_id: Optional - assess single student
        force_regenerate: If True, delete and recreate existing assessments

    Returns:
        Number of assessments created
    """

    # Determine which students to assess
    query = Student.query.filter_by(is_active=True)

    if student_id:
        query = query.filter_by(id=student_id)
    elif stream_id:
        query = query.filter_by(stream_id=stream_id)
    elif class_id:
        query = query.filter_by(class_id=class_id)

    students = query.all()
    assessments_created = 0

    for student in students:
        # Get all applicable fee items for this student
        applicable_fees = get_applicable_fees_for_student(student, term, year)

        for fee_item, rate_info in applicable_fees:
            # Check if assessment already exists
            existing = FeeAssessment.query.filter_by(
                student_id=student.id,
                fee_item_id=fee_item.id,
                term=term,
                year=year
            ).first()

            if existing:
                if not force_regenerate:
                    continue  # Skip if already assessed and not forcing regenerate
                else:
                    # Delete existing assessment if forcing regenerate
                    db.session.delete(existing)
                    db.session.flush()  # Ensure deletion is processed

            # Calculate amount
            amount = calculate_fee_amount(student, fee_item, rate_info)

            if amount > 0:
                assessment = FeeAssessment(
                    student_id=student.id,
                    fee_item_id=fee_item.id,
                    term=term,
                    year=year,
                    description=f"{fee_item.name} - Term {term} {year}",
                    amount=amount,
                    base_rate=rate_info.get('base_rate'),
                    quantity=rate_info.get('quantity', 1)
                )

                db.session.add(assessment)
                assessments_created += 1

    db.session.commit()
    return assessments_created

def get_rate_info(fee_rate, student, fee_item):
    """Extract rate information for calculation"""
    if fee_item.is_per_km and student.transport_distance_km and fee_rate.rate_per_km:
        return {
            'source': 'standard_per_km',
            'base_rate': fee_rate.rate_per_km,
            'quantity': student.transport_distance_km
        }
    elif fee_rate.amount:
        return {
            'source': 'standard_fixed',
            'base_rate': fee_rate.amount,
            'quantity': 1
        }
    return None


def calculate_fee_amount(student, fee_item, rate_info):
    """Calculate the actual fee amount"""
    base_rate = rate_info.get('base_rate', 0)
    quantity = rate_info.get('quantity', 1)

    return Decimal(str(base_rate)) * Decimal(str(quantity))

@app.route('/fees/individual/<int:student_id>', methods=['GET', 'POST'])
@login_required
def individual_fee_assignment(student_id):
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        assignment = StudentFeeAssignment(
            student_id=student_id,
            fee_item_id=int(request.form['fee_item_id']),
            term=int(request.form['term']),
            year=int(request.form['year']),
            custom_amount=Decimal(request.form['custom_amount']) if request.form['custom_amount'] else None,
            custom_rate_per_km=Decimal(request.form['custom_rate_per_km']) if request.form[
                'custom_rate_per_km'] else None,
            custom_distance=Decimal(request.form['custom_distance']) if request.form['custom_distance'] else None,
            notes=request.form['notes'],
            assigned_by=current_user.id
        )

        db.session.add(assignment)
        db.session.commit()
        flash('Individual fee assignment created', 'success')
        return redirect(url_for('student_detail', student_id=student_id))

    fee_items = FeeItem.query.filter_by(is_active=True).all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    existing_assignments = StudentFeeAssignment.query.filter_by(
        student_id=student_id,
        is_active=True
    ).all()

    return render_template('fees/individual_assignment.html',
                           student=student,
                           fee_items=fee_items,
                           current_year=current_year,
                           existing_assignments=existing_assignments)


# ===========================
#  PAYMENT MANAGEMENT
# ===========================
@app.route('/payments')
@login_required
def payment_list():
    page = request.args.get('page', 1, type=int)
    per_page = 20

    search = request.args.get('search', '')
    payment_mode = request.args.get('payment_mode', '')
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    query = Payment.query

    if search:
        query = query.join(Student).filter(or_(
            Payment.receipt_number.contains(search),
            Student.admission_no.contains(search),
            Student.first_name.contains(search),
            Student.last_name.contains(search)
        ))

    if payment_mode:
        query = query.filter_by(payment_mode=PaymentMode(payment_mode))

    if from_date:
        query = query.filter(Payment.payment_date >= datetime.strptime(from_date, '%Y-%m-%d').date())

    if to_date:
        query = query.filter(Payment.payment_date <= datetime.strptime(to_date, '%Y-%m-%d').date())

    payments = query.order_by(desc(Payment.created_at)) \
        .paginate(page=page, per_page=per_page, error_out=False)

    return render_template('payments/list.html', payments=payments, search=search)


@app.route('/payments/add/<int:student_id>', methods=['GET', 'POST'])
@login_required
def add_payment(student_id):
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        # Generate receipt number
        last_receipt = Payment.query.order_by(desc(Payment.id)).first()
        receipt_no = f"RCT{(last_receipt.id + 1):06d}" if last_receipt else "RCT000001"

        payment = Payment(
            student_id=student_id,
            amount=Decimal(request.form['amount']),
            payment_date=datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date(),
            payment_mode=PaymentMode(request.form['payment_mode']),
            receipt_number=receipt_no,
            mpesa_code=request.form.get('mpesa_code'),
            bank_slip_number=request.form.get('bank_slip_number'),
            cheque_number=request.form.get('cheque_number'),
            notes=request.form.get('notes'),
            processed_by=current_user.id
        )

        db.session.add(payment)
        db.session.commit()

        flash(f'Payment {receipt_no} recorded successfully', 'success')
        return redirect(url_for('allocate_payment', payment_id=payment.id))

    # Get outstanding assessments with proper calculation
    outstanding_assessments = []
    assessments = FeeAssessment.query.filter_by(student_id=student_id).all()

    for assessment in assessments:
        # Calculate how much has been paid for this assessment
        paid_amount = db.session.query(func.sum(PaymentAllocation.amount)) \
                          .filter_by(assessment_id=assessment.id).scalar() or Decimal('0')

        outstanding = assessment.amount - paid_amount

        # Only include if there's an outstanding balance
        if outstanding > 0:
            outstanding_assessments.append({
                'assessment': assessment,
                'outstanding': outstanding
            })

    return render_template('payments/add.html',
                           student=student,
                           outstanding_assessments=outstanding_assessments,
                           today_date=date.today().isoformat())  # Add this line

@app.route('/payments/<int:payment_id>/allocate', methods=['GET', 'POST'])
@login_required
def allocate_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)

    if request.method == 'POST':
        try:
            # Get all allocation inputs
            allocation_inputs = request.form.getlist('allocations')
            total_allocated = Decimal('0')
            allocations_to_create = []

            for allocation_data in allocation_inputs:
                # Skip empty values
                if not allocation_data or not allocation_data.strip():
                    continue

                # Check if the format is correct (assessment_id:amount)
                if ':' not in allocation_data:
                    continue

                try:
                    parts = allocation_data.split(':')
                    if len(parts) != 2:
                        continue

                    assessment_id = int(parts[0])
                    amount = Decimal(parts[1])

                    # Skip zero or negative amounts
                    if amount <= 0:
                        continue

                    # Verify the assessment exists and belongs to this student
                    assessment = FeeAssessment.query.get(assessment_id)
                    if not assessment or assessment.student_id != payment.student_id:
                        flash(f'Invalid assessment ID: {assessment_id}', 'error')
                        continue

                    # Check if amount exceeds outstanding balance for this assessment
                    already_allocated = db.session.query(func.sum(PaymentAllocation.amount)) \
                                            .filter_by(assessment_id=assessment_id).scalar() or Decimal('0')
                    outstanding = assessment.amount - already_allocated

                    if amount > outstanding:
                        flash(f'Amount for {assessment.fee_item.name} exceeds outstanding balance', 'warning')
                        amount = outstanding

                    allocations_to_create.append({
                        'assessment_id': assessment_id,
                        'amount': amount
                    })
                    total_allocated += amount

                except (ValueError, TypeError) as e:
                    # Skip malformed allocation data
                    continue

            # Validate total allocation doesn't exceed payment amount
            if total_allocated > payment.amount:
                flash('Total allocation exceeds payment amount', 'error')
                db.session.rollback()
            elif total_allocated == 0:
                flash('No valid allocations provided', 'warning')
            else:
                # Create all allocations
                for alloc_data in allocations_to_create:
                    allocation = PaymentAllocation(
                        payment_id=payment_id,
                        assessment_id=alloc_data['assessment_id'],
                        amount=alloc_data['amount']
                    )
                    db.session.add(allocation)

                db.session.commit()
                # Format currency manually instead of using |currency filter
                flash(f'Payment allocated successfully. Total: KSh {total_allocated:,.2f}', 'success')
                return redirect(url_for('student_detail', student_id=payment.student_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error allocating payment: {str(e)}', 'error')

    # Get outstanding assessments for GET request or failed POST
    outstanding_assessments = db.session.query(
        FeeAssessment,
        (FeeAssessment.amount - func.coalesce(func.sum(PaymentAllocation.amount), 0)).label('outstanding')
    ).filter_by(student_id=payment.student_id) \
        .outerjoin(PaymentAllocation) \
        .group_by(FeeAssessment.id) \
        .having(func.coalesce(func.sum(PaymentAllocation.amount), 0) < FeeAssessment.amount) \
        .all()

    return render_template('payments/allocate.html',
                           payment=payment,
                           outstanding_assessments=outstanding_assessments)


@app.route('/payments/<int:payment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(payment_id):
    """Edit an existing payment record"""
    payment = Payment.query.get_or_404(payment_id)

    if request.method == 'POST':
        try:
            # Store old values for audit log
            old_amount = payment.amount
            old_date = payment.payment_date
            old_mode = payment.payment_mode.value

            # Update payment details
            payment.amount = Decimal(request.form['amount'])
            payment.payment_date = datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date()
            payment.payment_mode = PaymentMode(request.form['payment_mode'])

            # Update payment method specific fields
            payment.mpesa_code = request.form.get('mpesa_code') if payment.payment_mode == PaymentMode.MPESA else None
            payment.bank_slip_number = request.form.get(
                'bank_slip_number') if payment.payment_mode == PaymentMode.BANK else None
            payment.cheque_number = request.form.get(
                'cheque_number') if payment.payment_mode == PaymentMode.CHEQUE else None

            # Update notes
            edit_reason = request.form.get('edit_reason', '')
            existing_notes = payment.notes or ''
            audit_note = f"\n\n[EDITED on {datetime.now().strftime('%Y-%m-%d %H:%M')} by {current_user.name}]\nReason: {edit_reason}\nChanges: Amount {old_amount} -> {payment.amount}, Date {old_date} -> {payment.payment_date}, Mode {old_mode} -> {payment.payment_mode.value}"
            payment.notes = existing_notes + audit_note

            db.session.commit()

            flash(f'Payment {payment.receipt_number} updated successfully', 'success')
            return redirect(url_for('payment_detail', payment_id=payment_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating payment: {str(e)}', 'error')

    return render_template('payments/edit.html', payment=payment)


@app.route('/payments/<int:payment_id>/delete', methods=['POST'])
@login_required
def delete_payment(payment_id):
    """Delete a payment record (with restrictions)"""
    payment = Payment.query.get_or_404(payment_id)

    # Check if payment has allocations
    if payment.allocations:
        flash('Cannot delete payment - it has been allocated to fee assessments. Remove allocations first.', 'error')
        return redirect(url_for('payment_detail', payment_id=payment_id))

    # Require admin role for deletion
    if current_user.role != UserRole.ADMIN:
        flash('Only administrators can delete payment records', 'error')
        return redirect(url_for('payment_detail', payment_id=payment_id))

    try:
        receipt_number = payment.receipt_number
        student_id = payment.student_id

        db.session.delete(payment)
        db.session.commit()

        flash(f'Payment {receipt_number} deleted successfully', 'success')
        return redirect(url_for('student_payments', student_id=student_id))

    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting payment: {str(e)}', 'error')
        return redirect(url_for('payment_detail', payment_id=payment_id))

# ===========================
#  REPORTING
# ===========================
@app.route('/reports')
@login_required
def reports():
    """
    Reports & Analytics dashboard.
    Renders the reports/index.html template with model access for counts and stats.
    """
    return render_template(
        'reports/index.html',
        now=datetime.now(),
        Student=Student,
        Payment=Payment,
        Vehicle=Vehicle,
        Expense=Expense
    )


@app.route('/reports/class_summary')
@login_required
def class_summary_report():
    term = request.args.get('term', type=int)
    year = request.args.get('year', type=int)
    class_id = request.args.get('class_id', type=int)

    current_year = AcademicYear.query.filter_by(is_current=True).first()
    if not term and current_year:
        year = current_year.year
        term = 1  # Default to term 1

    # Get class summary data
    classes = Class.query.all()
    summary_data = []

    target_classes = [Class.query.get(class_id)] if class_id else classes

    for cls in target_classes:
        if not cls:
            continue

        students = Student.query.filter_by(class_id=cls.id, is_active=True).all()

        class_summary = {
            'class': cls,
            'total_students': len(students),
            'total_assessed': 0,
            'total_paid': 0,
            'outstanding': 0,
            'streams': {}
        }

        for student in students:
            # Get student assessments for this term/year
            assessments = FeeAssessment.query.filter_by(
                student_id=student.id,
                term=term,
                year=year
            ).all() if term and year else []

            student_assessed = sum(a.amount for a in assessments)

            # Get student payments allocated to these assessments
            student_paid = db.session.query(func.sum(PaymentAllocation.amount)) \
                               .join(FeeAssessment) \
                               .filter(FeeAssessment.student_id == student.id,
                                       FeeAssessment.term == term,
                                       FeeAssessment.year == year) \
                               .scalar() or 0

            class_summary['total_assessed'] += student_assessed
            class_summary['total_paid'] += student_paid

            # Stream breakdown
            stream_key = student.stream.name if student.stream else 'No Stream'
            if stream_key not in class_summary['streams']:
                class_summary['streams'][stream_key] = {
                    'students': 0, 'assessed': 0, 'paid': 0
                }

            class_summary['streams'][stream_key]['students'] += 1
            class_summary['streams'][stream_key]['assessed'] += student_assessed
            class_summary['streams'][stream_key]['paid'] += student_paid

        class_summary['outstanding'] = class_summary['total_assessed'] - class_summary['total_paid']
        summary_data.append(class_summary)

    return render_template('reports/class_summary.html',
                           summary_data=summary_data,
                           classes=classes,
                           selected_class=class_id,
                           term=term,
                           year=year)


@app.route('/reports/student_statement/<int:student_id>')
@login_required
def student_statement(student_id):
    student = Student.query.get_or_404(student_id)
    balance_summary = get_student_balance_summary(student_id)

    # Get detailed transaction history
    assessments = FeeAssessment.query.filter_by(student_id=student_id) \
        .order_by(FeeAssessment.year, FeeAssessment.term) \
        .all()

    payments = Payment.query.filter_by(student_id=student_id) \
        .order_by(Payment.payment_date) \
        .all()

    return render_template('reports/student_statement.html',
                           student=student,
                           balance_summary=balance_summary,
                           assessments=assessments,
                           payments=payments)


@app.route('/reports/outstanding_fees')
@login_required
def outstanding_fees_report():
    class_id = request.args.get('class_id', type=int)

    query = Student.query.filter_by(is_active=True)
    if class_id:
        query = query.filter_by(class_id=class_id)

    students_with_outstanding = []

    for student in query.all():
        balance = student.get_current_balance()
        if balance > 0:
            students_with_outstanding.append({
                'student': student,
                'balance': balance
            })

    # Sort by balance descending
    students_with_outstanding.sort(key=lambda x: x['balance'], reverse=True)

    classes = Class.query.all()

    return render_template('reports/outstanding_fees.html',
                           students_with_outstanding=students_with_outstanding,
                           classes=classes,
                           selected_class=class_id)


@app.route('/reports/vehicle_revenue')
@login_required
def vehicle_revenue_report():
    term = request.args.get('term', type=int)
    year = request.args.get('year', type=int)

    current_year = AcademicYear.query.filter_by(is_current=True).first()
    if not term and current_year:
        year = current_year.year
        term = 1

    vehicles = Vehicle.query.filter_by(is_active=True).all()
    vehicle_data = []

    # Get transport fee item
    transport_fee = FeeItem.query.filter_by(code='TRANSPORT', is_active=True).first()

    # Get transport rate for this term/year
    rate = None
    if transport_fee and term and year:
        rate = FeeRate.query.filter_by(
            fee_item_id=transport_fee.id,
            term=term,
            year=year,
            is_active=True
        ).first()

    for vehicle in vehicles:
        students = Student.query.filter_by(
            vehicle_id=vehicle.id,
            is_active=True
        ).all()

        vehicle_summary = {
            'vehicle': vehicle,
            'total_students': len(students),
            'total_distance': sum(float(s.transport_distance_km or 0) for s in students),
            'total_assessed': Decimal('0'),
            'total_paid': Decimal('0'),
            'total_balance': Decimal('0'),
            'students': []
        }

        for student in students:
            if not student.transport_distance_km:
                continue

            # Calculate assessed amount
            assessed_amount = Decimal('0')
            if rate and rate.rate_per_km:
                assessed_amount = Decimal(str(student.transport_distance_km)) * rate.rate_per_km

            # Get actual payments for this student's transport fees
            paid_amount = Decimal('0')
            if transport_fee and term and year:
                # Get transport assessments for this student in this term
                transport_assessments = FeeAssessment.query.filter_by(
                    student_id=student.id,
                    fee_item_id=transport_fee.id,
                    term=term,
                    year=year
                ).all()

                # Sum up payments allocated to these assessments
                for assessment in transport_assessments:
                    allocated = db.session.query(func.sum(PaymentAllocation.amount)) \
                                    .filter_by(assessment_id=assessment.id) \
                                    .scalar() or 0
                    paid_amount += Decimal(str(allocated))

            balance = assessed_amount - paid_amount

            vehicle_summary['total_assessed'] += assessed_amount
            vehicle_summary['total_paid'] += paid_amount
            vehicle_summary['total_balance'] += balance

            vehicle_summary['students'].append({
                'student': student,
                'distance': float(student.transport_distance_km),
                'assessed': assessed_amount,
                'paid': paid_amount,
                'balance': balance
            })

        vehicle_data.append(vehicle_summary)

    # Calculate totals
    totals = {
        'vehicles': len(vehicle_data),
        'students': sum(v['total_students'] for v in vehicle_data),
        'distance': sum(v['total_distance'] for v in vehicle_data),
        'assessed': sum(v['total_assessed'] for v in vehicle_data),
        'paid': sum(v['total_paid'] for v in vehicle_data),
        'balance': sum(v['total_balance'] for v in vehicle_data)
    }

    return render_template('reports/vehicle_revenue.html',
                           vehicle_data=vehicle_data,
                           totals=totals,
                           term=term,
                           year=year,
                           rate=rate,
                           transport_fee=transport_fee)

# ===========================
#  EXPENSE MANAGEMENT
# ===========================
@app.route('/expenses')
@login_required
def expense_list():
    from datetime import datetime, timedelta
    from sqlalchemy import func, extract

    page = request.args.get('page', 1, type=int)
    per_page = 20

    category_id = request.args.get('category_id', type=int)

    query = Expense.query
    if category_id:
        query = query.filter_by(category_id=category_id)

    expenses = query.order_by(desc(Expense.expense_date)) \
        .paginate(page=page, per_page=per_page, error_out=False)

    categories = ExpenseCategory.query.filter_by(is_active=True).all()

    # Calculate summary statistics
    current_date = datetime.now()
    current_month = current_date.month
    current_year = current_date.year

    # Base query for filtered calculations
    base_query = Expense.query
    if category_id:
        base_query = base_query.filter_by(category_id=category_id)

    # This month's expenses
    this_month_total = base_query.filter(
        extract('month', Expense.expense_date) == current_month,
        extract('year', Expense.expense_date) == current_year
    ).with_entities(func.sum(Expense.amount)).scalar() or 0

    # This year's expenses
    this_year_total = base_query.filter(
        extract('year', Expense.expense_date) == current_year
    ).with_entities(func.sum(Expense.amount)).scalar() or 0

    # Calculate average per month (based on months this year that have passed)
    months_passed = current_month
    average_per_month = this_year_total / months_passed if months_passed > 0 and this_year_total > 0 else 0

    # Total expenses (all time or filtered)
    total_expenses = base_query.with_entities(func.sum(Expense.amount)).scalar() or 0

    # Calculate category breakdown for current month
    category_breakdown = db.session.query(
        ExpenseCategory.name,
        ExpenseCategory.code,
        func.sum(Expense.amount).label('total')
    ).join(Expense).filter(
        extract('month', Expense.expense_date) == current_month,
        extract('year', Expense.expense_date) == current_year
    ).group_by(ExpenseCategory.id).all()

    # Get top spending category this month
    top_category = category_breakdown[0] if category_breakdown else None

    # Calculate percentage change from last month
    last_month = current_month - 1 if current_month > 1 else 12
    last_month_year = current_year if current_month > 1 else current_year - 1

    last_month_total = base_query.filter(
        extract('month', Expense.expense_date) == last_month,
        extract('year', Expense.expense_date) == last_month_year
    ).with_entities(func.sum(Expense.amount)).scalar() or 0

    # Calculate percentage change
    if last_month_total > 0:
        month_change_percent = ((float(this_month_total) - float(last_month_total)) / float(last_month_total)) * 100
    else:
        month_change_percent = 100 if this_month_total > 0 else 0

    return render_template('expenses/list.html',
                           expenses=expenses,
                           categories=categories,
                           selected_category=category_id,
                           this_month_total=float(this_month_total),
                           average_per_month=float(average_per_month),
                           total_expenses=float(total_expenses),
                           category_breakdown=category_breakdown,
                           top_category=top_category,
                           month_change_percent=month_change_percent,
                           current_month_name=current_date.strftime('%B'),
                           current_year=current_year)

@app.route('/expenses/add', methods=['GET', 'POST'])
@login_required
def add_expense():
    if request.method == 'POST':
        expense = Expense(
            category_id=int(request.form['category_id']),
            description=request.form['description'],
            amount=Decimal(request.form['amount']),
            expense_date=datetime.strptime(request.form['expense_date'], '%Y-%m-%d').date(),
            payment_method=PaymentMode(request.form['payment_method']) if request.form['payment_method'] else None,
            reference_number=request.form.get('reference_number'),
            supplier_name=request.form.get('supplier_name'),
            approved_by=request.form.get('approved_by'),
            notes=request.form.get('notes'),
            created_by=current_user.id
        )

        db.session.add(expense)
        db.session.commit()
        flash('Expense recorded successfully', 'success')
        return redirect(url_for('expense_list'))

    categories = ExpenseCategory.query.filter_by(is_active=True).all()
    return render_template('expenses/add.html', categories=categories)


@app.route('/expenses/category/add', methods=['GET', 'POST'])
@login_required
def add_expense_category():
    if request.method == 'POST':
        category = ExpenseCategory(
            name=request.form['name'],
            code=request.form['code'].upper(),
            description=request.form.get('description')
        )

        db.session.add(category)
        db.session.commit()
        flash('Expense category added successfully', 'success')
        return redirect(url_for('expense_list'))

    return render_template('expenses/add_category.html')


# ===========================
#  PROMOTION MANAGEMENT
# ===========================
@app.route('/promotions')
@login_required
def promotion_management():
    academic_years = AcademicYear.query.order_by(desc(AcademicYear.year)).all()
    recent_promotions = StudentPromotion.query.order_by(desc(StudentPromotion.promotion_date)).limit(20).all()

    return render_template('promotions/index.html',
                           academic_years=academic_years,
                           recent_promotions=recent_promotions)


@app.route('/promotions/bulk', methods=['GET', 'POST'])
@login_required
def bulk_promotion():
    if request.method == 'POST':
        from_class_id = int(request.form['from_class_id'])
        academic_year = int(request.form['academic_year'])
        stream_id = int(request.form['stream_id']) if request.form['stream_id'] else None

        # Get students to promote
        query = Student.query.filter_by(class_id=from_class_id, is_active=True)
        if stream_id:
            query = query.filter_by(stream_id=stream_id)

        students = query.all()
        promoted_count = 0

        for student in students:
            if student.promote_to_next_class(academic_year, PromotionStatus.PROMOTED):
                promoted_count += 1

        db.session.commit()
        flash(f'Successfully promoted {promoted_count} students', 'success')
        return redirect(url_for('promotion_management'))

    classes = Class.query.all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('promotions/bulk.html',
                           classes=classes,
                           current_year=current_year)


@app.route('/promotions/individual/<int:student_id>', methods=['GET', 'POST'])
@login_required
def individual_promotion(student_id):
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        promotion = StudentPromotion(
            student_id=student_id,
            from_class_id=student.class_id,
            from_stream_id=student.stream_id,
            to_class_id=int(request.form['to_class_id']),
            to_stream_id=int(request.form['to_stream_id']) if request.form['to_stream_id'] else None,
            academic_year=int(request.form['academic_year']),
            status=PromotionStatus(request.form['status']),
            notes=request.form.get('notes'),
            processed_by=current_user.id
        )

        # Update student's current class
        student.class_id = promotion.to_class_id
        student.stream_id = promotion.to_stream_id

        db.session.add(promotion)
        db.session.commit()

        flash('Student promotion processed successfully', 'success')
        return redirect(url_for('student_detail', student_id=student_id))

    classes = Class.query.all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('promotions/individual.html',
                           student=student,
                           classes=classes,
                           current_year=current_year)


# ===========================
#  API ENDPOINTS
# ===========================
@app.route('/api/student_search')
@login_required
def student_search():
    query = request.args.get('q', '')
    if len(query) < 2:
        return jsonify([])

    students = Student.query.filter(
        Student.is_active == True,
        or_(
            Student.admission_no.contains(query),
            Student.first_name.contains(query),
            Student.last_name.contains(query)
        )
    ).limit(10).all()

    return jsonify([{
        'id': s.id,
        'admission_no': s.admission_no,
        'name': s.full_name,
        'class': f"{s.class_obj.name}{'-' + s.stream.name if s.stream else ''}"
    } for s in students])


@app.route('/api/student_balance/<int:student_id>')
@login_required
def get_student_balance(student_id):
    student = Student.query.get_or_404(student_id)
    balance = student.get_current_balance()

    return jsonify({
        'student_id': student_id,
        'admission_no': student.admission_no,
        'name': student.full_name,
        'balance': float(balance)
    })


# ===========================
#  ERROR HANDLERS
# ===========================
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('errors/500.html'), 500


# ===========================
#  DATABASE INITIALIZATION
# ===========================
def init_database():
    """Initialize database with default data"""
    db.create_all()

    # Create default data if needed
    if not ExpenseCategory.query.first():
        default_categories = [
            ('ADMIN', 'Administration'),
            ('MAINT', 'Maintenance'),
            ('UTIL', 'Utilities'),
            ('TRANS', 'Transport'),
            ('FOOD', 'Food & Catering'),
            ('SUPP', 'Supplies'),
            ('OTHER', 'Other')
        ]

        for code, name in default_categories:
            category = ExpenseCategory(code=code, name=name)
            db.session.add(category)

        db.session.commit()

    # Create default fee items if needed
    if not FeeItem.query.first():
        default_fees = [
            ('TUITION', 'Tuition Fee', FeeScope.CLASS_LEVEL, False),
            ('BOARDING', 'Boarding Fee', FeeScope.CLASS_LEVEL, False),
            ('TRANSPORT', 'Transport Fee', FeeScope.INDIVIDUAL, True),
            ('MEALS', 'Meals Fee', FeeScope.CLASS_LEVEL, False),
            ('UNIFORM', 'Uniform Fee', FeeScope.UNIVERSAL, False),
            ('BOOKS', 'Books & Materials', FeeScope.CLASS_LEVEL, False),
            ('ACTIVITY', 'Activity Fee', FeeScope.UNIVERSAL, False),
            ('SWIMMING', 'Swimming Fee', FeeScope.INDIVIDUAL, False),
            ('EXAM', 'Examination Fee', FeeScope.CLASS_LEVEL, False)
        ]

        for code, name, scope, is_per_km in default_fees:
            fee = FeeItem(code=code, name=name, scope=scope, is_per_km=is_per_km)
            db.session.add(fee)

        db.session.commit()


# ===========================
#  TEMPLATE FILTERS
# ===========================
@app.template_filter('currency')
def currency_filter(amount):
    """Format amount as currency"""
    if amount is None:
        return "KSh 0.00"
    return f"KSh {amount:,.2f}"


@app.template_filter('percentage')
def percentage_filter(value):
    """Format value as percentage"""
    if value is None:
        return "0%"
    return f"{value:.1f}%"


# Add these routes to your main Flask application file

# ===========================
#  EDIT/DELETE CLASS ROUTES
# ===========================
@app.route('/academic/class/edit/<int:class_id>', methods=['GET', 'POST'])
@login_required
def edit_class(class_id):
    class_obj = Class.query.get_or_404(class_id)

    if request.method == 'POST':
        class_obj.name = request.form['name']
        class_obj.level = request.form['level']
        class_obj.next_class_id = request.form['next_class_id'] if request.form['next_class_id'] else None

        db.session.commit()
        flash('Class updated successfully', 'success')
        return redirect(url_for('academic_management'))

    classes = Class.query.filter(Class.id != class_id).all()  # Exclude current class from next_class options
    return render_template('academic/edit_class.html', class_obj=class_obj, classes=classes)


@app.route('/academic/class/delete/<int:class_id>', methods=['POST'])
@login_required
def delete_class(class_id):
    class_obj = Class.query.get_or_404(class_id)

    # Check if class has students
    student_count = Student.query.filter_by(class_id=class_id, is_active=True).count()
    if student_count > 0:
        flash(f'Cannot delete class {class_obj.name} - it has {student_count} active students', 'error')
        return redirect(url_for('academic_management'))

    # Check if class has streams
    stream_count = Stream.query.filter_by(class_id=class_id).count()
    if stream_count > 0:
        flash(f'Cannot delete class {class_obj.name} - it has {stream_count} streams. Delete streams first.', 'error')
        return redirect(url_for('academic_management'))

    # Check if class is referenced as next_class by other classes
    referencing_classes = Class.query.filter_by(next_class_id=class_id).all()
    if referencing_classes:
        class_names = ', '.join([c.name for c in referencing_classes])
        flash(f'Cannot delete class {class_obj.name} - it is set as the next class for: {class_names}', 'error')
        return redirect(url_for('academic_management'))

    db.session.delete(class_obj)
    db.session.commit()
    flash(f'Class {class_obj.name} deleted successfully', 'success')
    return redirect(url_for('academic_management'))


# ===========================
#  EDIT/DELETE STREAM ROUTES
# ===========================
@app.route('/academic/stream/edit/<int:stream_id>', methods=['GET', 'POST'])
@login_required
def edit_stream(stream_id):
    stream = Stream.query.get_or_404(stream_id)

    if request.method == 'POST':
        stream.name = request.form['name']
        stream.capacity = int(request.form['capacity']) if request.form['capacity'] else 40

        db.session.commit()
        flash('Stream updated successfully', 'success')
        return redirect(url_for('academic_management'))

    return render_template('academic/edit_stream.html', stream=stream)


@app.route('/academic/stream/delete/<int:stream_id>', methods=['POST'])
@login_required
def delete_stream(stream_id):
    stream = Stream.query.get_or_404(stream_id)

    # Check if stream has students
    student_count = Student.query.filter_by(stream_id=stream_id, is_active=True).count()
    if student_count > 0:
        flash(f'Cannot delete stream {stream.class_obj.name}-{stream.name} - it has {student_count} active students',
              'error')
        return redirect(url_for('academic_management'))

    class_name = stream.class_obj.name
    stream_name = stream.name

    db.session.delete(stream)
    db.session.commit()
    flash(f'Stream {class_name}-{stream_name} deleted successfully', 'success')
    return redirect(url_for('academic_management'))


# ===========================
#  EDIT/DELETE ACADEMIC YEAR ROUTES
# ===========================
@app.route('/academic/year/edit/<int:year_id>', methods=['GET', 'POST'])
@login_required
def edit_academic_year(year_id):
    academic_year = AcademicYear.query.get_or_404(year_id)

    if request.method == 'POST':
        academic_year.year = int(request.form['year'])
        academic_year.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        academic_year.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()

        # Handle current year setting
        is_current = bool(request.form.get('is_current'))
        if is_current and not academic_year.is_current:
            # Unset other current years
            AcademicYear.query.update({'is_current': False})
            academic_year.is_current = True
        elif not is_current and academic_year.is_current:
            academic_year.is_current = False

        db.session.commit()
        flash('Academic year updated successfully', 'success')
        return redirect(url_for('academic_management'))

    return render_template('academic/edit_year.html', academic_year=academic_year)


@app.route('/academic/year/delete/<int:year_id>', methods=['POST'])
@login_required
def delete_academic_year(year_id):
    academic_year = AcademicYear.query.get_or_404(year_id)

    # Check if year has fee assessments
    assessment_count = FeeAssessment.query.filter_by(year=academic_year.year).count()
    if assessment_count > 0:
        flash(f'Cannot delete academic year {academic_year.year} - it has {assessment_count} fee assessments', 'error')
        return redirect(url_for('academic_management'))

    # Check if year has fee rates
    rate_count = FeeRate.query.filter_by(year=academic_year.year).count()
    if rate_count > 0:
        flash(f'Cannot delete academic year {academic_year.year} - it has {rate_count} fee rates', 'error')
        return redirect(url_for('academic_management'))

    # Check if year has student promotions
    promotion_count = StudentPromotion.query.filter_by(academic_year=academic_year.year).count()
    if promotion_count > 0:
        flash(f'Cannot delete academic year {academic_year.year} - it has {promotion_count} student promotions',
              'error')
        return redirect(url_for('academic_management'))

    year_value = academic_year.year
    db.session.delete(academic_year)
    db.session.commit()
    flash(f'Academic year {year_value} deleted successfully', 'success')
    return redirect(url_for('academic_management'))


# ===========================
#  STUDENT EDIT/DELETE ROUTES
# ===========================
@app.route('/students/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        student.admission_no = request.form['admission_no']
        student.first_name = request.form['first_name']
        student.last_name = request.form['last_name']
        student.date_of_birth = datetime.strptime(request.form['date_of_birth'], '%Y-%m-%d').date() if request.form[
            'date_of_birth'] else None
        student.class_id = int(request.form['class_id'])
        student.stream_id = int(request.form['stream_id']) if request.form['stream_id'] else None
        student.student_type = StudentType(request.form['student_type'])
        student.parent_name = request.form['parent_name']
        student.parent_phone = request.form['parent_phone']
        student.parent_email = request.form['parent_email']
        student.vehicle_id = int(request.form['vehicle_id']) if request.form['vehicle_id'] else None
        student.transport_distance_km = Decimal(request.form['transport_distance_km']) if request.form[
            'transport_distance_km'] else None

        db.session.commit()
        flash('Student updated successfully', 'success')
        return redirect(url_for('student_detail', student_id=student_id))

    classes = Class.query.all()
    streams = Stream.query.filter_by(class_id=student.class_id).all()
    vehicles = Vehicle.query.filter_by(is_active=True).all()

    return render_template('students/edit.html',
                           student=student,
                           classes=classes,
                           streams=streams,
                           vehicles=vehicles)


@app.route('/students/delete/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)

    # Check if student has payments
    payment_count = Payment.query.filter_by(student_id=student_id).count()
    if payment_count > 0:
        flash(
            f'Cannot delete student {student.full_name} - they have {payment_count} payment records. Consider deactivating instead.',
            'error')
        return redirect(url_for('student_detail', student_id=student_id))

    # Check if student has fee assessments
    assessment_count = FeeAssessment.query.filter_by(student_id=student_id).count()
    if assessment_count > 0:
        flash(
            f'Cannot delete student {student.full_name} - they have {assessment_count} fee assessments. Consider deactivating instead.',
            'error')
        return redirect(url_for('student_detail', student_id=student_id))

    # Check if student has promotions
    promotion_count = StudentPromotion.query.filter_by(student_id=student_id).count()
    if promotion_count > 0:
        flash(
            f'Cannot delete student {student.full_name} - they have promotion records. Consider deactivating instead.',
            'error')
        return redirect(url_for('student_detail', student_id=student_id))

    # Check if student has individual fee assignments
    assignment_count = StudentFeeAssignment.query.filter_by(student_id=student_id).count()
    if assignment_count > 0:
        flash(
            f'Cannot delete student {student.full_name} - they have individual fee assignments. Consider deactivating instead.',
            'error')
        return redirect(url_for('student_detail', student_id=student_id))

    student_name = student.full_name
    admission_no = student.admission_no

    db.session.delete(student)
    db.session.commit()
    flash(f'Student {student_name} ({admission_no}) deleted successfully', 'success')
    return redirect(url_for('student_list'))


@app.route('/students/deactivate/<int:student_id>', methods=['POST'])
@login_required
def deactivate_student(student_id):
    student = Student.query.get_or_404(student_id)

    student.is_active = False
    student.deactivation_date = date.today()

    db.session.commit()
    flash(f'Student {student.full_name} deactivated successfully', 'success')
    return redirect(url_for('student_detail', student_id=student_id))


@app.route('/students/reactivate/<int:student_id>', methods=['POST'])
@login_required
def reactivate_student(student_id):
    student = Student.query.get_or_404(student_id)

    student.is_active = True
    student.deactivation_date = None

    db.session.commit()
    flash(f'Student {student.full_name} reactivated successfully', 'success')
    return redirect(url_for('student_detail', student_id=student_id))


@app.template_global()
def today():
    return date.today()


# Add these routes to your Flask application (app.py)

# ===========================
#  FEE ITEM EDIT/DELETE ROUTES
# ===========================
@app.route('/fees/item/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
def edit_fee_item(item_id):
    fee_item = FeeItem.query.get_or_404(item_id)

    # Get usage statistics
    existing_rates_count = FeeRate.query.filter_by(fee_item_id=item_id).count()
    usage_stats = None

    if existing_rates_count > 0:
        total_assessments = FeeAssessment.query.filter_by(fee_item_id=item_id).count()
        total_assignments = StudentFeeAssignment.query.filter_by(fee_item_id=item_id).count()
        total_revenue = db.session.query(func.sum(FeeAssessment.amount)).filter_by(fee_item_id=item_id).scalar() or 0

        usage_stats = {
            'total_rates': existing_rates_count,
            'total_assessments': total_assessments,
            'total_assignments': total_assignments,
            'total_revenue': total_revenue
        }

    if request.method == 'POST':
        fee_item.name = request.form['name']
        fee_item.code = request.form['code'].upper()
        fee_item.description = request.form['description']
        fee_item.scope = FeeScope(request.form['scope'])
        fee_item.is_per_km = bool(request.form.get('is_per_km'))

        db.session.commit()
        flash('Fee item updated successfully', 'success')
        return redirect(url_for('fee_management'))

    return render_template('fees/edit_item.html',
                           fee_item=fee_item,
                           existing_rates_count=existing_rates_count,
                           usage_stats=usage_stats)


@app.route('/fees/item/delete/<int:item_id>', methods=['POST'])
@login_required
def delete_fee_item(item_id):
    fee_item = FeeItem.query.get_or_404(item_id)

    # Check if fee item has assessments
    assessment_count = FeeAssessment.query.filter_by(fee_item_id=item_id).count()
    if assessment_count > 0:
        flash(f'Cannot delete fee item {fee_item.name} - it has {assessment_count} assessments', 'error')
        return redirect(url_for('fee_management'))

    # Check if fee item has rates
    rate_count = FeeRate.query.filter_by(fee_item_id=item_id).count()
    if rate_count > 0:
        flash(f'Cannot delete fee item {fee_item.name} - it has {rate_count} fee rates', 'error')
        return redirect(url_for('fee_management'))

    # Check if fee item has individual assignments
    assignment_count = StudentFeeAssignment.query.filter_by(fee_item_id=item_id).count()
    if assignment_count > 0:
        flash(f'Cannot delete fee item {fee_item.name} - it has {assignment_count} individual assignments', 'error')
        return redirect(url_for('fee_management'))

    fee_item_name = fee_item.name
    db.session.delete(fee_item)
    db.session.commit()
    flash(f'Fee item {fee_item_name} deleted successfully', 'success')
    return redirect(url_for('fee_management'))


@app.route('/fees/item/activate/<int:item_id>', methods=['POST'])
@login_required
def activate_fee_item(item_id):
    fee_item = FeeItem.query.get_or_404(item_id)
    fee_item.is_active = True
    db.session.commit()
    flash(f'Fee item {fee_item.name} activated successfully', 'success')
    return redirect(url_for('fee_management'))


@app.route('/fees/item/deactivate/<int:item_id>', methods=['POST'])
@login_required
def deactivate_fee_item(item_id):
    fee_item = FeeItem.query.get_or_404(item_id)
    fee_item.is_active = False
    db.session.commit()
    flash(f'Fee item {fee_item.name} deactivated successfully', 'success')
    return redirect(url_for('fee_management'))


# ===========================
#  FEE RATE EDIT/DELETE ROUTES
# ===========================
@app.route('/fees/rate/edit/<int:rate_id>', methods=['GET', 'POST'])
@login_required
def edit_fee_rate(rate_id):
    fee_rate = FeeRate.query.get_or_404(rate_id)

    # Get existing assessments count
    existing_assessments_count = FeeAssessment.query.filter_by(fee_item_id=fee_rate.fee_item_id).count()

    # Get usage statistics
    usage_stats = None
    assessments_with_rate = db.session.query(FeeAssessment) \
        .filter_by(fee_item_id=fee_rate.fee_item_id) \
        .filter_by(term=fee_rate.term, year=fee_rate.year).all()

    if assessments_with_rate:
        students_affected = len(set(a.student_id for a in assessments_with_rate))
        total_amount = sum(a.amount for a in assessments_with_rate)

        usage_stats = {
            'assessments_count': len(assessments_with_rate),
            'students_affected': students_affected,
            'total_amount': total_amount
        }

    if request.method == 'POST':
        fee_rate.fee_item_id = int(request.form['fee_item_id'])
        fee_rate.term = int(request.form['term'])
        fee_rate.year = int(request.form['year'])
        fee_rate.class_id = int(request.form['class_id']) if request.form['class_id'] else None
        fee_rate.stream_id = int(request.form['stream_id']) if request.form['stream_id'] else None
        fee_rate.student_type = StudentType(request.form['student_type']) if request.form['student_type'] else None
        fee_rate.amount = Decimal(request.form['amount']) if request.form['amount'] else None
        fee_rate.rate_per_km = Decimal(request.form['rate_per_km']) if request.form['rate_per_km'] else None
        fee_rate.is_active = bool(request.form.get('is_active'))

        db.session.commit()
        flash('Fee rate updated successfully', 'success')
        return redirect(url_for('fee_management'))

    fee_items = FeeItem.query.filter_by(is_active=True).all()
    classes = Class.query.all()

    return render_template('fees/edit_rate.html',
                           fee_rate=fee_rate,
                           fee_items=fee_items,
                           classes=classes,
                           existing_assessments_count=existing_assessments_count,
                           usage_stats=usage_stats)


@app.route('/fees/rate/delete/<int:rate_id>', methods=['POST'])
@login_required
def delete_fee_rate(rate_id):
    fee_rate = FeeRate.query.get_or_404(rate_id)

    # Check if rate has been used in assessments
    assessment_count = db.session.query(FeeAssessment) \
        .filter_by(fee_item_id=fee_rate.fee_item_id) \
        .filter_by(term=fee_rate.term, year=fee_rate.year).count()

    if assessment_count > 0:
        flash(
            f'Cannot delete fee rate - it has been used in {assessment_count} assessments. Consider deactivating instead.',
            'error')
        return redirect(url_for('fee_management'))

    fee_item_name = fee_rate.fee_item.name
    term_year = f"Term {fee_rate.term}/{fee_rate.year}"

    db.session.delete(fee_rate)
    db.session.commit()
    flash(f'Fee rate for {fee_item_name} ({term_year}) deleted successfully', 'success')
    return redirect(url_for('fee_management'))


# ===========================
#  STUDENT ASSIGNMENT ROUTES
# ===========================
@app.route('/fees/assignment/edit/<int:assignment_id>', methods=['GET', 'POST'])
@login_required
def edit_student_assignment(assignment_id):
    assignment = StudentFeeAssignment.query.get_or_404(assignment_id)
    student = assignment.student

    if request.method == 'POST':
        assignment.fee_item_id = int(request.form['fee_item_id'])
        assignment.term = int(request.form['term'])
        assignment.year = int(request.form['year'])
        assignment.custom_amount = Decimal(request.form['custom_amount']) if request.form['custom_amount'] else None
        assignment.custom_rate_per_km = Decimal(request.form['custom_rate_per_km']) if request.form[
            'custom_rate_per_km'] else None
        assignment.custom_distance = Decimal(request.form['custom_distance']) if request.form[
            'custom_distance'] else None
        assignment.notes = request.form['notes']
        assignment.is_active = bool(request.form.get('is_active', True))

        db.session.commit()
        flash('Individual fee assignment updated successfully', 'success')
        return redirect(url_for('student_detail', student_id=student.id))

    fee_items = FeeItem.query.filter_by(is_active=True).all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('fees/edit_assignment.html',
                           assignment=assignment,
                           student=student,
                           fee_items=fee_items,
                           current_year=current_year)


@app.route('/fees/assignment/delete/<int:assignment_id>', methods=['POST'])
@login_required
def delete_student_assignment(assignment_id):
    assignment = StudentFeeAssignment.query.get_or_404(assignment_id)
    student_id = assignment.student_id

    # Check if assignment has been used in assessments
    assessment_count = db.session.query(FeeAssessment) \
        .filter_by(student_id=assignment.student_id) \
        .filter_by(fee_item_id=assignment.fee_item_id) \
        .filter_by(term=assignment.term, year=assignment.year).count()

    if assessment_count > 0:
        flash('Cannot delete assignment - assessments have been generated based on it. Consider deactivating instead.',
              'error')
        return redirect(url_for('student_detail', student_id=student_id))

    db.session.delete(assignment)
    db.session.commit()
    flash('Individual fee assignment deleted successfully', 'success')
    return redirect(url_for('student_detail', student_id=student_id))


# ===========================
#  API ENDPOINTS FOR FEE MANAGEMENT
# ===========================
@app.route('/api/get_standard_rate', methods=['POST'])
@login_required
def get_standard_rate():
    """Get standard rate information for a fee item, term, year, and student"""
    try:
        fee_item_id = int(request.form['fee_item_id'])
        term = int(request.form['term'])
        year = int(request.form['year'])
        student_id = int(request.form['student_id'])

        student = Student.query.get_or_404(student_id)
        fee_item = FeeItem.query.get_or_404(fee_item_id)

        # Find applicable rate
        rate = None
        scope_description = ""

        if fee_item.scope == FeeScope.UNIVERSAL:
            rate = FeeRate.query.filter_by(
                fee_item_id=fee_item_id,
                term=term,
                year=year,
                class_id=None,
                stream_id=None,
                is_active=True
            ).first()
            scope_description = "All students"

        elif fee_item.scope == FeeScope.STREAM_LEVEL and student.stream_id:
            rate = FeeRate.query.filter_by(
                fee_item_id=fee_item_id,
                term=term,
                year=year,
                stream_id=student.stream_id,
                student_type=student.student_type,
                is_active=True
            ).first()

            if not rate:
                rate = FeeRate.query.filter_by(
                    fee_item_id=fee_item_id,
                    term=term,
                    year=year,
                    stream_id=student.stream_id,
                    student_type=None,
                    is_active=True
                ).first()

            scope_description = f"{student.class_obj.name}-{student.stream.name}"

        elif fee_item.scope == FeeScope.CLASS_LEVEL:
            rate = FeeRate.query.filter_by(
                fee_item_id=fee_item_id,
                term=term,
                year=year,
                class_id=student.class_id,
                student_type=student.student_type,
                is_active=True
            ).first()

            if not rate:
                rate = FeeRate.query.filter_by(
                    fee_item_id=fee_item_id,
                    term=term,
                    year=year,
                    class_id=student.class_id,
                    student_type=None,
                    is_active=True
                ).first()

            scope_description = student.class_obj.name

        if rate:
            rate_info = {
                'amount': f"KSh {rate.amount:,.2f}" if rate.amount else None,
                'rate_per_km': f"KSh {rate.rate_per_km:,.2f}" if rate.rate_per_km else None,
                'scope_description': scope_description,
                'student_distance': float(student.transport_distance_km) if student.transport_distance_km else None,
                'calculated_amount': None
            }

            if rate.rate_per_km and student.transport_distance_km:
                calculated = float(rate.rate_per_km * student.transport_distance_km)
                rate_info['calculated_amount'] = f"KSh {calculated:,.2f}"

            return jsonify({'success': True, 'rate': rate_info})
        else:
            return jsonify({'success': False, 'message': 'No standard rate found'})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/assess_fees', methods=['GET', 'POST'])
@login_required
def assess_fees_with_preview():
    """Enhanced assess fees with preview functionality"""
    if request.method == 'POST':
        # Check if this is a preview request
        if request.form.get('preview'):
            return handle_fee_assessment_preview()
        else:
            # Actual assessment generation
            term = int(request.form['term'])
            year = int(request.form['year'])
            assessment_scope = request.form['assessment_scope']

            class_id = None
            stream_id = None
            student_id = None

            if assessment_scope == 'class':
                class_id = int(request.form['class_id']) if request.form['class_id'] else None
            elif assessment_scope == 'stream':
                class_id = int(request.form['class_id']) if request.form['class_id'] else None
                stream_id = int(request.form['stream_id']) if request.form['stream_id'] else None
            elif assessment_scope == 'individual':
                student_id = int(request.form['student_id']) if request.form['student_id'] else None

            # Check if this is a dry run
            dry_run = bool(request.form.get('dry_run'))

            if not dry_run:
                assessments_created = generate_fee_assessments(
                    term=term,
                    year=year,
                    class_id=class_id,
                    stream_id=stream_id,
                    student_id=student_id
                )

                flash(f'Generated {assessments_created} fee assessments', 'success')
            else:
                flash('Dry run completed - no assessments were created', 'info')

            return redirect(url_for('fee_management'))

    classes = Class.query.all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('fees/assess.html',
                           classes=classes,
                           current_year=current_year)


def handle_fee_assessment_preview():
    """Handle preview request for fee assessments"""
    try:
        term = int(request.form['term'])
        year = int(request.form['year'])
        assessment_scope = request.form['assessment_scope']
        skip_existing = bool(request.form.get('skip_existing'))
        include_transport = bool(request.form.get('include_transport'))

        # Determine target students
        query = Student.query.filter_by(is_active=True)

        if assessment_scope == 'class' and request.form.get('class_id'):
            query = query.filter_by(class_id=int(request.form['class_id']))
        elif assessment_scope == 'stream' and request.form.get('stream_id'):
            query = query.filter_by(stream_id=int(request.form['stream_id']))
        elif assessment_scope == 'individual' and request.form.get('student_id'):
            query = query.filter_by(id=int(request.form['student_id']))

        students = query.all()

        # Simulate assessment generation
        preview_data = {
            'students_count': len(students),
            'assessments_count': 0,
            'total_amount': 'KSh 0.00',
            'skipped_count': 0,
            'fee_breakdown': [],
            'warnings': []
        }

        total_amount = Decimal('0')
        fee_totals = {}

        for student in students:
            # Check existing assessments if skip_existing is enabled
            if skip_existing:
                existing = FeeAssessment.query.filter_by(
                    student_id=student.id,
                    term=term,
                    year=year
                ).first()

                if existing:
                    preview_data['skipped_count'] += 1
                    continue

            # Get applicable fees for this student
            applicable_fees = get_applicable_fees_for_student(student, term, year)

            for fee_item, rate_info in applicable_fees:
                # Skip transport fees if not included
                if not include_transport and fee_item.code == 'TRANSPORT':
                    continue

                amount = calculate_fee_amount(student, fee_item, rate_info)

                if amount > 0:
                    preview_data['assessments_count'] += 1
                    total_amount += amount

                    # Track fee breakdown
                    if fee_item.code not in fee_totals:
                        fee_totals[fee_item.code] = {
                            'name': fee_item.name,
                            'code': fee_item.code,
                            'students_count': 0,
                            'amount': Decimal('0')
                        }

                    fee_totals[fee_item.code]['students_count'] += 1
                    fee_totals[fee_item.code]['amount'] += amount

        # Format fee breakdown
        for code, data in fee_totals.items():
            preview_data['fee_breakdown'].append({
                'code': data['code'],
                'name': data['name'],
                'students_count': data['students_count'],
                'amount': f"KSh {data['amount']:,.2f}"
            })

        preview_data['total_amount'] = f"KSh {total_amount:,.2f}"

        # Add warnings
        if preview_data['assessments_count'] == 0:
            preview_data['warnings'].append('No new assessments will be created with the current criteria.')

        if not include_transport:
            transport_students = sum(1 for s in students if s.vehicle_id)
            if transport_students > 0:
                preview_data['warnings'].append(
                    f'{transport_students} students have transport assignments but transport fees are excluded.')

        return jsonify({'success': True, 'data': preview_data})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Preview failed: {str(e)}'})


# Add these routes to your app.py file

# ===========================
#  ADDITIONAL FEE ROUTES
# ===========================

@app.route('/fees/items')
@login_required
def fee_items_list():
    """Display all fee items with filtering"""
    status = request.args.get('status', '')
    scope = request.args.get('scope', '')

    query = FeeItem.query

    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)

    if scope:
        query = query.filter_by(scope=FeeScope(scope))

    fee_items = query.all()

    return render_template('fees/items_list.html', fee_items=fee_items)


@app.route('/students/<int:student_id>/payments')
@login_required
def student_payments(student_id):
    """Display all payments for a specific student"""
    student = Student.query.get_or_404(student_id)
    balance_summary = get_student_balance_summary(student_id)

    # Get all payments ordered by date
    payments = Payment.query.filter_by(student_id=student_id) \
        .order_by(desc(Payment.payment_date)).all()

    return render_template('payments/student_payments.html',
                           student=student,
                           payments=payments,
                           balance_summary=balance_summary)



def amount_to_words(amount):
    """Convert amount to words for receipt"""
    amount = float(amount)
    shillings = int(amount)
    cents = int((amount - shillings) * 100)

    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen',
             'Nineteen']

    def convert_below_thousand(n):
        if n == 0:
            return ''
        elif n < 10:
            return ones[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (' ' + ones[n % 10] if n % 10 != 0 else '')
        else:
            return ones[n // 100] + ' Hundred' + (' ' + convert_below_thousand(n % 100) if n % 100 != 0 else '')

    if shillings == 0:
        words = 'Zero Shillings'
    else:
        words = ''

        # Millions
        if shillings >= 1000000:
            words += convert_below_thousand(shillings // 1000000) + ' Million '
            shillings %= 1000000

        # Thousands
        if shillings >= 1000:
            words += convert_below_thousand(shillings // 1000) + ' Thousand '
            shillings %= 1000

        # Remaining
        if shillings > 0:
            words += convert_below_thousand(shillings)

        words = words.strip() + ' Shillings'

    if cents > 0:
        words += ' and ' + convert_below_thousand(cents) + ' Cents'

    return words



@app.route('/fees/rates')
@login_required
def fee_rates_list():
    """Display all fee rates with filtering"""
    term = request.args.get('term', type=int)
    year = request.args.get('year', type=int)
    fee_item_id = request.args.get('fee_item_id', type=int)
    class_id = request.args.get('class_id', type=int)
    status = request.args.get('status', '')

    query = FeeRate.query

    if term:
        query = query.filter_by(term=term)
    if year:
        query = query.filter_by(year=year)
    if fee_item_id:
        query = query.filter_by(fee_item_id=fee_item_id)
    if class_id:
        query = query.filter_by(class_id=class_id)
    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)

    current_rates = query.all()
    fee_items = FeeItem.query.filter_by(is_active=True).all()
    classes = Class.query.all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    return render_template('fees/rates_list.html',
                           current_rates=current_rates,
                           fee_items=fee_items,
                           classes=classes,
                           current_year=current_year)


from flask import make_response
from io import BytesIO
from reportlab.lib.pagesizes import A5
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Circle, String
from decimal import Decimal


def amount_to_words(amount):
    """Convert amount to words"""
    amount = float(amount)
    shillings = int(amount)
    cents = int((amount - shillings) * 100)

    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen',
             'Nineteen']

    def convert_below_thousand(n):
        if n == 0:
            return ''
        elif n < 10:
            return ones[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (' ' + ones[n % 10] if n % 10 != 0 else '')
        else:
            return ones[n // 100] + ' Hundred' + (' ' + convert_below_thousand(n % 100) if n % 100 != 0 else '')

    if shillings == 0:
        words = 'Zero'
    else:
        words = ''
        if shillings >= 1000000:
            words += convert_below_thousand(shillings // 1000000) + ' Million '
            shillings %= 1000000
        if shillings >= 1000:
            words += convert_below_thousand(shillings // 1000) + ' Thousand '
            shillings %= 1000
        if shillings > 0:
            words += convert_below_thousand(shillings)
        words = words.strip()

    if cents > 0:
        words += ' and ' + convert_below_thousand(cents) + ' Cents'

    return words


def generate_receipt_pdf(payment):
    """Generate simple PDF receipt matching the image style"""
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=A5,
                            rightMargin=10 * mm, leftMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    school_name_style = ParagraphStyle(
        'SchoolName',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=3,
        fontName='Helvetica-Bold'
    )

    school_info_style = ParagraphStyle(
        'SchoolInfo',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=2
    )

    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading2'],
        fontSize=14,
        alignment=TA_CENTER,
        spaceAfter=8,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )

    # Header with logo
    header_table_data = [['']]
    header_table = Table(header_table_data, colWidths=[128 * mm])

    # Logo drawing
    logo_drawing = Drawing(128 * mm, 25 * mm)
    logo_drawing.add(Circle(20 * mm, 12.5 * mm, 8 * mm,
                            fillColor=colors.lightgrey,
                            strokeColor=colors.black,
                            strokeWidth=2))
    logo_drawing.add(String(20 * mm, 11 * mm, 'SLM',
                            fontSize=14,
                            fillColor=colors.black,
                            textAnchor='middle',
                            fontName='Helvetica-Bold'))
    elements.append(logo_drawing)

    # School name and details
    elements.append(Paragraph("ST LUKE MOGOBICH ACADEMY", school_name_style))
    elements.append(Paragraph("P.O. BOX 1234-00100, Nairobi.", school_info_style))
    elements.append(Paragraph("Tel: 0706 950 210 / 0701 693 009", school_info_style))
    elements.append(Paragraph("Email: stlukemogobich@gmail.com", school_info_style))
    elements.append(Paragraph("<b>Chokaa- Utawala Road</b>", school_info_style))

    # Horizontal line
    line_table = Table([['']], colWidths=[128 * mm], rowHeights=[2])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, colors.black),
    ]))
    elements.append(Spacer(1, 3 * mm))
    elements.append(line_table)

    # Title
    elements.append(Paragraph("OFFICIAL RECEIPT", title_style))

    # Receipt details
    receipt_data = [
        ['Receipt No:', payment.receipt_number, 'Rect Date:', payment.payment_date.strftime('%d.%B.%Y')],
        ['Received From:', payment.student.full_name.upper(), '', ''],
        ['ADM NO:', payment.student.admission_no, 'Class:',
         f"{payment.student.class_obj.name}{' ' + payment.student.stream.name if payment.student.stream else ''}"]
    ]

    receipt_table = Table(receipt_data, colWidths=[28 * mm, 40 * mm, 24 * mm, 36 * mm])
    receipt_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
        ('FONTNAME', (2, 2), (2, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(receipt_table)
    elements.append(Spacer(1, 5 * mm))

    # Amount in words
    amount_para = Paragraph(f"<b>Amount..</b> {amount_to_words(payment.amount)} Shillings",
                            ParagraphStyle('Amount', fontSize=10, spaceAfter=5))
    elements.append(amount_para)
    elements.append(Spacer(1, 3 * mm))

    # Fee breakdown table
    fee_data = [['VOTEHEAD', 'AMOUNT']]

    if payment.allocations:
        for allocation in payment.allocations:
            fee_data.append([
                allocation.assessment.fee_item.name.upper(),
                f"{allocation.amount:.2f}"
            ])
    else:
        fee_data.append(['UNALLOCATED PAYMENT', f"{payment.amount:.2f}"])

    # Add prepayments row
    fee_data.append(['PREPAYMENTS', '0.00'])

    # Calculate balance
    total_assessed = db.session.query(db.func.sum(FeeAssessment.amount)) \
                         .filter_by(student_id=payment.student.id).scalar() or Decimal('0')

    total_paid = db.session.query(db.func.sum(PaymentAllocation.amount)) \
                     .join(FeeAssessment) \
                     .filter(FeeAssessment.student_id == payment.student.id).scalar() or Decimal('0')

    current_balance = Decimal(str(total_assessed)) - Decimal(str(total_paid))

    # Add arrears and total
    fee_data.append(['FEES ARREARS', f"{current_balance:.2f}"])
    fee_data.append(['TOTAL PAID:', f"{payment.amount:.2f}"])

    fee_table = Table(fee_data, colWidths=[90 * mm, 38 * mm])
    fee_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        # Last two rows bold
        ('BACKGROUND', (0, -2), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(fee_table)
    elements.append(Spacer(1, 5 * mm))

    # Footer details
    footer_data = [
        ['Balance:', f"{current_balance:.2f}",
         'BANK SLIP' if payment.payment_mode.value == 'BANK' else payment.payment_mode.value, ''],
        ['Payment Mode:',
         'BANK SLIP' if payment.payment_mode.value == 'BANK' else payment.payment_mode.value,
         payment.bank_slip_number if payment.bank_slip_number else
         payment.mpesa_code if payment.mpesa_code else
         payment.cheque_number if payment.cheque_number else '', ''],
        ['Received By:', payment.processor.name if payment.processor else 'SYSTEM',
         payment.payment_date.strftime('%d/%m/%Y'), '']
    ]

    footer_table = Table(footer_data, colWidths=[28 * mm, 36 * mm, 32 * mm, 32 * mm])
    footer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
    ]))
    elements.append(footer_table)
    elements.append(Spacer(1, 5 * mm))

    # Footer note
    line_table2 = Table([['']], colWidths=[128 * mm], rowHeights=[1])
    line_table2.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(line_table2)
    elements.append(Spacer(1, 2 * mm))

    footer_note = Paragraph(
        "<para align=center><font size=9>Fees once paid is neither Refundable nor Transferrable.</font></para>",
        styles['Normal']
    )
    elements.append(footer_note)

    doc.build(elements)
    buffer.seek(0)
    return buffer


@app.route('/payments/<int:payment_id>/download-pdf')
@login_required
def download_receipt_pdf(payment_id):
    """Download receipt as PDF"""
    payment = Payment.query.get_or_404(payment_id)

    try:
        pdf_buffer = generate_receipt_pdf(payment)

        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=Receipt_{payment.receipt_number}.pdf'

        return response
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('payment_detail', payment_id=payment_id))


# Add these routes to your app.py file (around line 1250 or after the payment_list route)

@app.route('/payments/<int:payment_id>')
@login_required
def payment_detail(payment_id):
    """Display detailed view of a single payment with receipt"""
    payment = Payment.query.get_or_404(payment_id)

    # Calculate correct balance
    total_assessed = db.session.query(db.func.sum(FeeAssessment.amount)) \
                         .filter_by(student_id=payment.student.id).scalar() or Decimal('0')

    total_paid = db.session.query(db.func.sum(PaymentAllocation.amount)) \
                     .join(FeeAssessment) \
                     .filter(FeeAssessment.student_id == payment.student.id).scalar() or Decimal('0')

    current_balance = Decimal(str(total_assessed)) - Decimal(str(total_paid))

    return render_template('payments/receipt.html',
                           payment=payment,
                           amount_to_words=amount_to_words,
                           total_assessed=total_assessed,
                           total_paid=total_paid,
                           current_balance=current_balance)


@app.route('/payments/<int:payment_id>/print')
@login_required
def print_receipt(payment_id):
    """Print-friendly view of receipt"""
    payment = Payment.query.get_or_404(payment_id)

    # Calculate correct balance
    total_assessed = db.session.query(db.func.sum(FeeAssessment.amount)) \
                         .filter_by(student_id=payment.student.id).scalar() or Decimal('0')

    total_paid = db.session.query(db.func.sum(PaymentAllocation.amount)) \
                     .join(FeeAssessment) \
                     .filter(FeeAssessment.student_id == payment.student.id).scalar() or Decimal('0')

    current_balance = Decimal(str(total_assessed)) - Decimal(str(total_paid))

    return render_template('payments/print_receipt.html',
                           payment=payment,
                           amount_to_words=amount_to_words,
                           total_assessed=total_assessed,
                           total_paid=total_paid,
                           current_balance=current_balance)


def amount_to_words(amount):
    """Convert amount to words for receipt"""
    amount = float(amount)
    shillings = int(amount)
    cents = int((amount - shillings) * 100)

    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen',
             'Nineteen']

    def convert_below_thousand(n):
        if n == 0:
            return ''
        elif n < 10:
            return ones[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (' ' + ones[n % 10] if n % 10 != 0 else '')
        else:
            return ones[n // 100] + ' Hundred' + (' ' + convert_below_thousand(n % 100) if n % 100 != 0 else '')

    if shillings == 0:
        words = 'Zero'
    else:
        words = ''
        if shillings >= 1000000:
            words += convert_below_thousand(shillings // 1000000) + ' Million '
            shillings %= 1000000
        if shillings >= 1000:
            words += convert_below_thousand(shillings // 1000) + ' Thousand '
            shillings %= 1000
        if shillings > 0:
            words += convert_below_thousand(shillings)
        words = words.strip()

    if cents > 0:
        words += ' and ' + convert_below_thousand(cents) + ' Cents'

    return words


# Also add the template filter at the end of your app.py (before if __name__ == '__main__':)
@app.template_filter('amount_to_words')
def amount_to_words_filter(amount):
    """Template filter for converting amounts to words"""
    return amount_to_words(amount)


# Add these imports at the top of your app.py
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file


# ===========================
#  EXCEL REPORT ROUTES
# ===========================

def create_excel_workbook():
    """Create a new Excel workbook with default styling"""
    wb = openpyxl.Workbook()
    return wb


def style_header_row(ws, row_num=1):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_adjust_column_width(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@app.route('/download/all-data')
@login_required
def download_all_data():
    """Download complete system data in Excel format"""
    wb = create_excel_workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["ST LUKE MOGOBICH ACADEMY - COMPLETE DATA EXPORT"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    ws_summary.append(["Category", "Count"])

    ws_summary.append(["Active Students", Student.query.filter_by(is_active=True).count()])
    ws_summary.append(["Total Payments", Payment.query.count()])
    ws_summary.append(["Active Vehicles", Vehicle.query.filter_by(is_active=True).count()])
    ws_summary.append(["Total Expenses", Expense.query.count()])
    ws_summary.append(["Classes", Class.query.count()])
    ws_summary.append(["Fee Items", FeeItem.query.count()])

    # Total Financial Summary
    total_assessed = db.session.query(func.sum(FeeAssessment.amount)).scalar() or 0
    total_paid = db.session.query(func.sum(Payment.amount)).scalar() or 0
    total_expenses = db.session.query(func.sum(Expense.amount)).scalar() or 0

    ws_summary.append([])
    ws_summary.append(["Financial Summary"])
    ws_summary.append(["Total Fees Assessed", f"KSh {float(total_assessed):,.2f}"])
    ws_summary.append(["Total Payments Received", f"KSh {float(total_paid):,.2f}"])
    ws_summary.append(["Outstanding Balance", f"KSh {float(total_assessed - total_paid):,.2f}"])
    ws_summary.append(["Total Expenses", f"KSh {float(total_expenses):,.2f}"])

    style_header_row(ws_summary, 4)
    style_header_row(ws_summary, 12)
    auto_adjust_column_width(ws_summary)

    # Students Sheet
    ws_students = wb.create_sheet("Students")
    ws_students.append(["Admission No", "First Name", "Last Name", "Class", "Stream",
                        "Student Type", "Parent Name", "Parent Phone", "Vehicle",
                        "Transport Distance (km)", "Status"])

    students = Student.query.all()
    for student in students:
        ws_students.append([
            student.admission_no,
            student.first_name,
            student.last_name,
            student.class_obj.name,
            student.stream.name if student.stream else "N/A",
            student.student_type.value,
            student.parent_name,
            student.parent_phone,
            student.vehicle.registration_number if student.vehicle else "N/A",
            float(student.transport_distance_km) if student.transport_distance_km else 0,
            "Active" if student.is_active else "Inactive"
        ])

    style_header_row(ws_students)
    auto_adjust_column_width(ws_students)

    # Payments Sheet
    ws_payments = wb.create_sheet("Payments")
    ws_payments.append(["Receipt No", "Date", "Student", "Admission No", "Amount",
                        "Payment Mode", "M-Pesa Code", "Bank Slip", "Processed By"])

    payments = Payment.query.order_by(desc(Payment.payment_date)).all()
    for payment in payments:
        ws_payments.append([
            payment.receipt_number,
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.student.full_name,
            payment.student.admission_no,
            float(payment.amount),
            payment.payment_mode.value,
            payment.mpesa_code or "",
            payment.bank_slip_number or "",
            payment.processor.name if payment.processor else "SYSTEM"
        ])

    style_header_row(ws_payments)
    auto_adjust_column_width(ws_payments)

    # Fee Assessments Sheet
    ws_assessments = wb.create_sheet("Fee Assessments")
    ws_assessments.append(["Student", "Admission No", "Fee Item", "Term", "Year",
                           "Amount", "Assessed Date"])

    assessments = FeeAssessment.query.order_by(desc(FeeAssessment.year),
                                               desc(FeeAssessment.term)).all()
    for assessment in assessments:
        ws_assessments.append([
            assessment.student.full_name,
            assessment.student.admission_no,
            assessment.fee_item.name,
            assessment.term,
            assessment.year,
            float(assessment.amount),
            assessment.assessed_date.strftime('%Y-%m-%d')
        ])

    style_header_row(ws_assessments)
    auto_adjust_column_width(ws_assessments)

    # Expenses Sheet
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(["Date", "Category", "Description", "Amount", "Payment Method",
                        "Reference", "Supplier", "Approved By"])

    expenses = Expense.query.order_by(desc(Expense.expense_date)).all()
    for expense in expenses:
        ws_expenses.append([
            expense.expense_date.strftime('%Y-%m-%d'),
            expense.category.name,
            expense.description,
            float(expense.amount),
            expense.payment_method.value if expense.payment_method else "N/A",
            expense.reference_number or "",
            expense.supplier_name or "",
            expense.approved_by or ""
        ])

    style_header_row(ws_expenses)
    auto_adjust_column_width(ws_expenses)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Complete_School_Data_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/students-by-class')
@login_required
def download_students_by_class():
    """Download students organized by class with separate tabs"""
    wb = create_excel_workbook()
    wb.remove(wb.active)

    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["CLASS SUMMARY REPORT"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    ws_summary.append(["Class", "Total Students", "Day Students", "Boarders", "Active", "Inactive"])

    classes = Class.query.all()
    total_students = 0

    for class_obj in classes:
        students = Student.query.filter_by(class_id=class_obj.id).all()
        active = len([s for s in students if s.is_active])
        inactive = len([s for s in students if not s.is_active])
        day = len([s for s in students if s.student_type == StudentType.DAY])
        boarders = len([s for s in students if s.student_type == StudentType.BOARDER])

        ws_summary.append([
            class_obj.name,
            len(students),
            day,
            boarders,
            active,
            inactive
        ])
        total_students += len(students)

    ws_summary.append([])
    ws_summary.append(["TOTAL", total_students])

    style_header_row(ws_summary, 4)
    auto_adjust_column_width(ws_summary)

    # Create sheet for each class
    for class_obj in classes:
        # Sanitize sheet name (Excel has 31 char limit and special char restrictions)
        sheet_name = class_obj.name[:31].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(sheet_name)

        ws.append([f"{class_obj.name} - Student List"])
        ws.append([])
        ws.append(["Admission No", "First Name", "Last Name", "Stream", "Type",
                   "Parent Name", "Parent Phone", "Vehicle", "Distance (km)", "Balance"])

        students = Student.query.filter_by(class_id=class_obj.id, is_active=True).all()

        for student in students:
            balance = student.get_current_balance()
            ws.append([
                student.admission_no,
                student.first_name,
                student.last_name,
                student.stream.name if student.stream else "N/A",
                student.student_type.value,
                student.parent_name,
                student.parent_phone,
                student.vehicle.registration_number if student.vehicle else "N/A",
                float(student.transport_distance_km) if student.transport_distance_km else 0,
                float(balance)
            ])

        # Add summary at bottom
        ws.append([])
        ws.append(["Total Students:", len(students)])

        style_header_row(ws, 3)
        auto_adjust_column_width(ws)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Students_By_Class_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/fee-collection-report')
@login_required
def download_fee_collection_report():
    """Download fee collection summary report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Fee Collection Summary"

    ws.append(["FEE COLLECTION SUMMARY REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Term", "Year", "Total Assessed", "Total Paid", "Outstanding"])

    # Get unique term/year combinations
    assessments = db.session.query(
        FeeAssessment.term,
        FeeAssessment.year,
        func.sum(FeeAssessment.amount).label('total_assessed')
    ).group_by(FeeAssessment.term, FeeAssessment.year) \
        .order_by(desc(FeeAssessment.year), desc(FeeAssessment.term)).all()

    for assessment in assessments:
        # Get total payments for this term/year
        total_paid = db.session.query(func.sum(PaymentAllocation.amount)) \
                         .join(FeeAssessment) \
                         .filter(FeeAssessment.term == assessment.term,
                                 FeeAssessment.year == assessment.year).scalar() or 0

        outstanding = float(assessment.total_assessed) - float(total_paid)

        ws.append([
            f"Term {assessment.term}",
            assessment.year,
            float(assessment.total_assessed),
            float(total_paid),
            outstanding
        ])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Fee_Collection_Report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/outstanding-report')
@login_required
def download_outstanding_report():
    """Download outstanding balances report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Outstanding Balances"

    ws.append(["OUTSTANDING BALANCES REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Admission No", "Student Name", "Class", "Parent Name", "Parent Phone", "Outstanding Balance"])

    students = Student.query.filter_by(is_active=True).all()
    students_with_balance = []

    for student in students:
        balance = student.get_current_balance()
        if balance > 0:
            students_with_balance.append({
                'student': student,
                'balance': balance
            })

    # Sort by balance descending
    students_with_balance.sort(key=lambda x: x['balance'], reverse=True)

    total_outstanding = 0
    for item in students_with_balance:
        student = item['student']
        balance = item['balance']
        total_outstanding += float(balance)

        ws.append([
            student.admission_no,
            student.full_name,
            f"{student.class_obj.name}{'-' + student.stream.name if student.stream else ''}",
            student.parent_name,
            student.parent_phone,
            float(balance)
        ])

    ws.append([])
    ws.append(["TOTAL OUTSTANDING:", "", "", "", "", total_outstanding])
    ws.append(["Number of Students:", len(students_with_balance)])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Outstanding_Balances_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/payment-history')
@login_required
def download_payment_history():
    """Download complete payment history"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Payment History"

    ws.append(["PAYMENT HISTORY REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Receipt No", "Date", "Student Name", "Admission No", "Class",
               "Amount", "Payment Mode", "M-Pesa/Bank Ref", "Processed By"])

    payments = Payment.query.order_by(desc(Payment.payment_date)).all()

    total_amount = 0
    for payment in payments:
        total_amount += float(payment.amount)

        ref = ""
        if payment.mpesa_code:
            ref = payment.mpesa_code
        elif payment.bank_slip_number:
            ref = payment.bank_slip_number
        elif payment.cheque_number:
            ref = payment.cheque_number

        ws.append([
            payment.receipt_number,
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.student.full_name,
            payment.student.admission_no,
            f"{payment.student.class_obj.name}{'-' + payment.student.stream.name if payment.student.stream else ''}",
            float(payment.amount),
            payment.payment_mode.value,
            ref,
            payment.processor.name if payment.processor else "SYSTEM"
        ])

    ws.append([])
    ws.append(["TOTAL PAYMENTS:", "", "", "", "", total_amount])
    ws.append(["Number of Transactions:", len(payments)])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Payment_History_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/expense-report')
@login_required
def download_expense_report():
    """Download expense report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws.append(["EXPENSE REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Date", "Category", "Description", "Amount", "Payment Method",
               "Reference", "Supplier", "Approved By"])

    expenses = Expense.query.order_by(desc(Expense.expense_date)).all()

    total_expenses = 0
    category_totals = {}

    for expense in expenses:
        total_expenses += float(expense.amount)

        category_name = expense.category.name
        if category_name not in category_totals:
            category_totals[category_name] = 0
        category_totals[category_name] += float(expense.amount)

        ws.append([
            expense.expense_date.strftime('%Y-%m-%d'),
            expense.category.name,
            expense.description,
            float(expense.amount),
            expense.payment_method.value if expense.payment_method else "N/A",
            expense.reference_number or "",
            expense.supplier_name or "",
            expense.approved_by or ""
        ])

    ws.append([])
    ws.append(["TOTAL EXPENSES:", "", "", total_expenses])
    ws.append([])
    ws.append(["BREAKDOWN BY CATEGORY:"])
    ws.append(["Category", "Amount"])

    for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
        ws.append([category, amount])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Expense_Report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/student-register')
@login_required
def download_student_register():
    """Download complete student register"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Student Register"

    ws.append(["COMPLETE STUDENT REGISTER"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Admission No", "First Name", "Last Name", "Date of Birth", "Class", "Stream",
               "Student Type", "Parent Name", "Parent Phone", "Parent Email",
               "Vehicle", "Transport Distance", "Admission Date", "Status"])

    students = Student.query.order_by(Student.admission_no).all()

    for student in students:
        ws.append([
            student.admission_no,
            student.first_name,
            student.last_name,
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else "N/A",
            student.class_obj.name,
            student.stream.name if student.stream else "N/A",
            student.student_type.value,
            student.parent_name,
            student.parent_phone,
            student.parent_email or "",
            student.vehicle.registration_number if student.vehicle else "N/A",
            float(student.transport_distance_km) if student.transport_distance_km else 0,
            student.admission_date.strftime('%Y-%m-%d') if student.admission_date else "N/A",
            "Active" if student.is_active else "Inactive"
        ])

    ws.append([])
    ws.append(["TOTAL STUDENTS:", len(students)])
    ws.append(["Active:", len([s for s in students if s.is_active])])
    ws.append(["Inactive:", len([s for s in students if not s.is_active])])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Student_Register_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/all-student-statements')
@login_required
def download_all_student_statements():
    """Download detailed fee statements for all students"""
    wb = create_excel_workbook()
    wb.remove(wb.active)

    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["STUDENT FEE STATEMENTS"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    ws_summary.append(["Admission No", "Student Name", "Class", "Total Assessed",
                       "Total Paid", "Balance"])

    students = Student.query.filter_by(is_active=True).order_by(Student.admission_no).all()

    grand_total_assessed = 0
    grand_total_paid = 0

    for student in students:
        # Calculate totals
        total_assessed = db.session.query(func.sum(FeeAssessment.amount)) \
                             .filter_by(student_id=student.id).scalar() or 0

        total_paid = db.session.query(func.sum(PaymentAllocation.amount)) \
                         .join(FeeAssessment) \
                         .filter(FeeAssessment.student_id == student.id).scalar() or 0

        balance = float(total_assessed) - float(total_paid)

        grand_total_assessed += float(total_assessed)
        grand_total_paid += float(total_paid)

        ws_summary.append([
            student.admission_no,
            student.full_name,
            f"{student.class_obj.name}{'-' + student.stream.name if student.stream else ''}",
            float(total_assessed),
            float(total_paid),
            balance
        ])

    ws_summary.append([])
    ws_summary.append(["TOTALS:", "", "", grand_total_assessed, grand_total_paid,
                       grand_total_assessed - grand_total_paid])

    style_header_row(ws_summary, 4)
    auto_adjust_column_width(ws_summary)

    # Create individual statement for each student (limit to first 50 to avoid huge files)
    for student in students[:50]:
        sheet_name = f"{student.admission_no}"[:31]
        ws = wb.create_sheet(sheet_name)

        ws.append([f"FEE STATEMENT - {student.full_name}"])
        ws.append([f"Admission No: {student.admission_no}"])
        ws.append([f"Class: {student.class_obj.name}{'-' + student.stream.name if student.stream else ''}"])
        ws.append([])

        # Assessments
        ws.append(["ASSESSMENTS"])
        ws.append(["Fee Item", "Term", "Year", "Amount", "Date"])

        assessments = FeeAssessment.query.filter_by(student_id=student.id) \
            .order_by(FeeAssessment.year, FeeAssessment.term).all()

        for assessment in assessments:
            ws.append([
                assessment.fee_item.name,
                assessment.term,
                assessment.year,
                float(assessment.amount),
                assessment.assessed_date.strftime('%Y-%m-%d')
            ])

        ws.append([])

        # Payments
        ws.append(["PAYMENTS"])
        ws.append(["Receipt No", "Date", "Amount", "Mode"])

        payments = Payment.query.filter_by(student_id=student.id) \
            .order_by(Payment.payment_date).all()

        for payment in payments:
            ws.append([
                payment.receipt_number,
                payment.payment_date.strftime('%Y-%m-%d'),
                float(payment.amount),
                payment.payment_mode.value
            ])

        ws.append([])

        # Summary
        total_assessed = sum(float(a.amount) for a in assessments)
        total_paid = sum(float(p.amount) for p in payments)

        ws.append(["SUMMARY"])
        ws.append(["Total Assessed:", total_assessed])
        ws.append(["Total Paid:", total_paid])
        ws.append(["Balance:", total_assessed - total_paid])

        style_header_row(ws, 6)
        style_header_row(ws, len(assessments) + 9)
        auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'All_Student_Statements_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/student-type-report')
@login_required
def download_student_type_report():
    """Download boarders vs day students analysis"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Student Type Analysis"

    ws.append(["BOARDERS VS DAY STUDENTS REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])

    # Overall summary
    total_students = Student.query.filter_by(is_active=True).count()
    day_students = Student.query.filter_by(is_active=True, student_type=StudentType.DAY).count()
    boarders = Student.query.filter_by(is_active=True, student_type=StudentType.BOARDER).count()

    ws.append(["OVERALL SUMMARY"])
    ws.append(["Category", "Count", "Percentage"])
    ws.append(["Total Students", total_students, "100%"])
    ws.append(
        ["Day Students", day_students, f"{(day_students / total_students * 100) if total_students > 0 else 0:.1f}%"])
    ws.append(["Boarders", boarders, f"{(boarders / total_students * 100) if total_students > 0 else 0:.1f}%"])
    ws.append([])

    # Breakdown by class
    ws.append(["BREAKDOWN BY CLASS"])
    ws.append(["Class", "Total", "Day Students", "Boarders"])

    classes = Class.query.all()
    for class_obj in classes:
        total = Student.query.filter_by(class_id=class_obj.id, is_active=True).count()
        day = Student.query.filter_by(class_id=class_obj.id, is_active=True,
                                      student_type=StudentType.DAY).count()
        board = Student.query.filter_by(class_id=class_obj.id, is_active=True,
                                        student_type=StudentType.BOARDER).count()

        ws.append([class_obj.name, total, day, board])

    style_header_row(ws, 5)
    style_header_row(ws, 11)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Student_Type_Analysis_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/vehicle-revenue')
@login_required
def download_vehicle_revenue():
    """Download vehicle revenue report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Vehicle Revenue"

    ws.append(["VEHICLE REVENUE REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Registration", "Make/Model", "Driver", "Students", "Total Distance",
               "Revenue Potential"])

    vehicles = Vehicle.query.filter_by(is_active=True).all()

    total_revenue = 0
    for vehicle in vehicles:
        students = Student.query.filter_by(vehicle_id=vehicle.id, is_active=True).all()
        total_distance = sum(float(s.transport_distance_km or 0) for s in students)

        # Calculate potential revenue (simplified - using current year/term 1)
        current_year = AcademicYear.query.filter_by(is_current=True).first()
        revenue = 0

        if current_year:
            transport_fee = FeeItem.query.filter_by(code='TRANSPORT').first()
            if transport_fee:
                rate = FeeRate.query.filter_by(
                    fee_item_id=transport_fee.id,
                    year=current_year.year,
                    term=1,
                    is_active=True
                ).first()

                if rate and rate.rate_per_km:
                    revenue = total_distance * float(rate.rate_per_km) * 3  # 3 terms

        total_revenue += revenue

        ws.append([
            vehicle.registration_number,
            f"{vehicle.make} {vehicle.model}",
            vehicle.driver_name,
            len(students),
            total_distance,
            revenue
        ])

    ws.append([])
    ws.append(["TOTAL REVENUE POTENTIAL:", "", "", "", "", total_revenue])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Vehicle_Revenue_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/transport-report')
@login_required
def download_transport_report():
    """Download transport assignments report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Transport Assignments"

    ws.append(["TRANSPORT ASSIGNMENTS REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Student Name", "Admission No", "Class", "Vehicle", "Driver",
               "Distance (km)", "Parent Contact"])

    students = Student.query.filter(Student.vehicle_id.isnot(None),
                                    Student.is_active == True).all()

    for student in students:
        ws.append([
            student.full_name,
            student.admission_no,
            f"{student.class_obj.name}{'-' + student.stream.name if student.stream else ''}",
            student.vehicle.registration_number,
            student.vehicle.driver_name,
            float(student.transport_distance_km) if student.transport_distance_km else 0,
            student.parent_phone
        ])

    ws.append([])
    ws.append(["TOTAL STUDENTS USING TRANSPORT:", len(students)])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Transport_Assignments_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/promotion-report')
@login_required
def download_promotion_report():
    """Download promotion history report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Promotion History"

    ws.append(["STUDENT PROMOTION HISTORY"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Student Name", "Admission No", "From Class", "From Stream",
               "To Class", "To Stream", "Academic Year", "Status", "Date"])

    promotions = StudentPromotion.query.order_by(desc(StudentPromotion.promotion_date)).all()

    for promotion in promotions:
        ws.append([
            promotion.student.full_name,
            promotion.student.admission_no,
            promotion.from_class.name if promotion.from_class else "N/A",
            promotion.from_stream.name if promotion.from_stream else "N/A",
            promotion.to_class.name,
            promotion.to_stream.name if promotion.to_stream else "N/A",
            promotion.academic_year,
            promotion.status.value,
            promotion.promotion_date.strftime('%Y-%m-%d')
        ])

    ws.append([])
    ws.append(["TOTAL PROMOTIONS:", len(promotions)])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Promotion_History_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/download/class-summary')
@login_required
def download_class_summary():
    """Download class summary report"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Class Summary"

    ws.append(["CLASS SUMMARY REPORT"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Class", "Streams", "Total Students", "Day Students", "Boarders",
               "With Transport", "Average Balance"])

    classes = Class.query.all()

    for class_obj in classes:
        streams = Stream.query.filter_by(class_id=class_obj.id).count()
        students = Student.query.filter_by(class_id=class_obj.id, is_active=True).all()

        day = len([s for s in students if s.student_type == StudentType.DAY])
        boarders = len([s for s in students if s.student_type == StudentType.BOARDER])
        with_transport = len([s for s in students if s.vehicle_id])

        # Calculate average balance
        total_balance = sum(float(s.get_current_balance()) for s in students)
        avg_balance = total_balance / len(students) if students else 0

        ws.append([
            class_obj.name,
            streams,
            len(students),
            day,
            boarders,
            with_transport,
            avg_balance
        ])

    style_header_row(ws, 4)
    auto_adjust_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Class_Summary_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/reports/fee-collection')
@login_required
def view_fee_collection_report():
    """View fee collection report in browser"""
    # Get unique term/year combinations
    assessments = db.session.query(
        FeeAssessment.term,
        FeeAssessment.year,
        func.sum(FeeAssessment.amount).label('total_assessed')
    ).group_by(FeeAssessment.term, FeeAssessment.year) \
        .order_by(desc(FeeAssessment.year), desc(FeeAssessment.term)).all()

    collection_data = []
    for assessment in assessments:
        total_paid = db.session.query(func.sum(PaymentAllocation.amount)) \
                         .join(FeeAssessment) \
                         .filter(FeeAssessment.term == assessment.term,
                                 FeeAssessment.year == assessment.year).scalar() or 0

        outstanding = float(assessment.total_assessed) - float(total_paid)

        collection_data.append({
            'term': assessment.term,
            'year': assessment.year,
            'assessed': float(assessment.total_assessed),
            'paid': float(total_paid),
            'outstanding': outstanding
        })

    return render_template('reports/fee_collection.html', collection_data=collection_data)


@app.route('/reports/custom-generate', methods=['POST'])
@login_required
def generate_custom_report():
    """Generate custom report based on user selection"""
    report_type = request.form.get('report_type')
    date_range = request.form.get('date_range')

    # Redirect to appropriate download route based on selection
    if report_type == 'students':
        return redirect(url_for('download_student_register'))
    elif report_type == 'payments':
        return redirect(url_for('download_payment_history'))
    elif report_type == 'assessments':
        return redirect(url_for('download_fee_collection_report'))
    elif report_type == 'expenses':
        return redirect(url_for('download_expense_report'))
    else:
        flash('Invalid report type selected', 'error')
        return redirect(url_for('reports'))


# Add these routes to your app.py file

# ===========================
#  EDIT EXPENSE ROUTE
# ===========================
@app.route('/expenses/edit/<int:expense_id>', methods=['GET', 'POST'])
@login_required
def edit_expense(expense_id):
    """Edit an existing expense record"""
    expense = Expense.query.get_or_404(expense_id)

    if request.method == 'POST':
        try:
            # Store old values for audit log
            old_amount = expense.amount
            old_category = expense.category.name
            old_date = expense.expense_date

            # Update expense details
            expense.category_id = int(request.form['category_id'])
            expense.expense_date = datetime.strptime(request.form['expense_date'], '%Y-%m-%d').date()
            expense.description = request.form['description']
            expense.amount = Decimal(request.form['amount'])
            expense.payment_method = PaymentMode(request.form['payment_method']) if request.form.get(
                'payment_method') else None
            expense.reference_number = request.form.get('reference_number')
            expense.supplier_name = request.form.get('supplier_name')
            expense.approved_by = request.form.get('approved_by')

            # Update notes with audit trail
            edit_reason = request.form.get('edit_reason', '')
            existing_notes = expense.notes or ''

            audit_note = f"\n\n[EDITED on {datetime.now().strftime('%Y-%m-%d %H:%M')} by {current_user.name}]\n"
            audit_note += f"Reason: {edit_reason}\n"
            audit_note += f"Changes: Amount {old_amount} -> {expense.amount}, "
            audit_note += f"Date {old_date} -> {expense.expense_date}, "
            audit_note += f"Category {old_category} -> {expense.category.name}"

            expense.notes = existing_notes + audit_note

            # Append new notes if provided
            new_notes = request.form.get('notes', '').strip()
            if new_notes and new_notes != existing_notes:
                expense.notes += f"\n{new_notes}"

            db.session.commit()

            flash(f'Expense updated successfully', 'success')
            return redirect(url_for('expense_list'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating expense: {str(e)}', 'error')

    categories = ExpenseCategory.query.filter_by(is_active=True).all()
    return render_template('expenses/edit.html', expense=expense, categories=categories)


@app.route('/expenses/delete/<int:expense_id>', methods=['POST'])
@login_required
def delete_expense(expense_id):
    """Delete an expense record (admin only)"""
    expense = Expense.query.get_or_404(expense_id)

    # Require admin role for deletion
    if current_user.role != UserRole.ADMIN:
        flash('Only administrators can delete expense records', 'error')
        return redirect(url_for('expense_list'))

    try:
        expense_desc = expense.description[:50]
        expense_amount = expense.amount

        db.session.delete(expense)
        db.session.commit()

        flash(f'Expense "{expense_desc}..." (KSh {expense_amount}) deleted successfully', 'success')
        return redirect(url_for('expense_list'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting expense: {str(e)}', 'error')
        return redirect(url_for('expense_list'))


# ===========================
#  EDIT VEHICLE ROUTE
# ===========================
@app.route('/vehicles/edit/<int:vehicle_id>', methods=['GET', 'POST'])
@login_required
def edit_vehicle(vehicle_id):
    """Edit an existing vehicle record"""
    vehicle = Vehicle.query.get_or_404(vehicle_id)

    if request.method == 'POST':
        try:
            # Check if registration number is being changed
            new_reg = request.form['registration_number'].upper()
            if new_reg != vehicle.registration_number:
                # Check if new registration number already exists
                existing = Vehicle.query.filter_by(registration_number=new_reg).first()
                if existing:
                    flash(f'Vehicle with registration number {new_reg} already exists', 'error')
                    return render_template('vehicles/edit.html', vehicle=vehicle)

            # Update vehicle details
            vehicle.registration_number = new_reg
            vehicle.make = request.form.get('make', '').strip()
            vehicle.model = request.form.get('model', '').strip()
            vehicle.capacity = int(request.form['capacity']) if request.form.get('capacity') else None
            vehicle.driver_name = request.form.get('driver_name', '').strip()
            vehicle.driver_phone = request.form.get('driver_phone', '').strip()
            vehicle.route_description = request.form.get('route_description', '').strip()
            vehicle.is_active = bool(request.form.get('is_active'))

            db.session.commit()

            flash(f'Vehicle {vehicle.registration_number} updated successfully', 'success')
            return redirect(url_for('vehicle_detail', vehicle_id=vehicle_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating vehicle: {str(e)}', 'error')

    return render_template('vehicles/edit.html', vehicle=vehicle)


@app.route('/vehicles/delete/<int:vehicle_id>', methods=['POST'])
@login_required
def delete_vehicle(vehicle_id):
    """Delete a vehicle record"""
    vehicle = Vehicle.query.get_or_404(vehicle_id)

    # Check if vehicle has students assigned
    student_count = Student.query.filter_by(vehicle_id=vehicle_id, is_active=True).count()
    if student_count > 0:
        flash(
            f'Cannot delete vehicle {vehicle.registration_number} - it has {student_count} active students assigned. Remove students first or deactivate the vehicle.',
            'error')
        return redirect(url_for('vehicle_detail', vehicle_id=vehicle_id))

    # Require admin role for deletion
    if current_user.role != UserRole.ADMIN:
        flash('Only administrators can delete vehicle records', 'error')
        return redirect(url_for('vehicle_detail', vehicle_id=vehicle_id))

    try:
        reg_number = vehicle.registration_number

        db.session.delete(vehicle)
        db.session.commit()

        flash(f'Vehicle {reg_number} deleted successfully', 'success')
        return redirect(url_for('vehicle_list'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting vehicle: {str(e)}', 'error')
        return redirect(url_for('vehicle_detail', vehicle_id=vehicle_id))


@app.route('/vehicles/deactivate/<int:vehicle_id>', methods=['POST'])
@login_required
def deactivate_vehicle(vehicle_id):
    """Deactivate a vehicle (soft delete)"""
    vehicle = Vehicle.query.get_or_404(vehicle_id)

    vehicle.is_active = False
    db.session.commit()

    flash(f'Vehicle {vehicle.registration_number} deactivated successfully', 'success')
    return redirect(url_for('vehicle_detail', vehicle_id=vehicle_id))


@app.route('/vehicles/reactivate/<int:vehicle_id>', methods=['POST'])
@login_required
def reactivate_vehicle(vehicle_id):
    """Reactivate a vehicle"""
    vehicle = Vehicle.query.get_or_404(vehicle_id)

    vehicle.is_active = True
    db.session.commit()

    flash(f'Vehicle {vehicle.registration_number} reactivated successfully', 'success')
    return redirect(url_for('vehicle_detail', vehicle_id=vehicle_id))


# Add these routes to your app.py file

# ===========================
#  USER MANAGEMENT ROUTES
# ===========================

@app.route('/users')
@login_required
def user_management():
    """View all users (Admin only)"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied. Only administrators can manage users.', 'error')
        return redirect(url_for('dashboard'))

    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('users/list.html', users=users)


@app.route('/api/user/<int:user_id>')
@login_required
def get_user_api(user_id):
    """API endpoint to get user data"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403

    user = User.query.get_or_404(user_id)
    return jsonify({
        'id': user.id,
        'name': user.name,
        'email': user.email,
        'role': user.role.value,
        'is_active': user.is_active,
        'google_id': user.google_id,
        'profile_pic': user.profile_pic,
        'created_at': user.created_at.strftime('%Y-%m-%d %H:%M'),
        'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else None
    })


@app.route('/users/edit/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    """Edit user details (Admin only)"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied. Only administrators can edit users.', 'error')
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(user_id)

    # Prevent editing your own account through this route
    if user.id == current_user.id:
        flash('You cannot edit your own account through this interface.', 'warning')
        return redirect(url_for('user_management'))

    try:
        # Update user details
        user.name = request.form['name']
        user.email = request.form['email']
        user.role = UserRole(request.form['role'])
        user.is_active = bool(request.form.get('is_active'))

        db.session.commit()

        flash(f'User {user.name} updated successfully', 'success')
        return redirect(url_for('user_management'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error updating user: {str(e)}', 'error')
        return redirect(url_for('user_management'))


@app.route('/users/deactivate/<int:user_id>', methods=['POST'])
@login_required
def deactivate_user(user_id):
    """Deactivate a user (Admin only)"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied. Only administrators can deactivate users.', 'error')
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(user_id)

    # Prevent deactivating yourself
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'warning')
        return redirect(url_for('user_management'))

    # Check if this is the last active admin
    if user.role == UserRole.ADMIN:
        active_admins = User.query.filter_by(role=UserRole.ADMIN, is_active=True).count()
        if active_admins <= 1:
            flash('Cannot deactivate the last active administrator.', 'error')
            return redirect(url_for('user_management'))

    user.is_active = False
    db.session.commit()

    flash(f'User {user.name} has been deactivated', 'success')
    return redirect(url_for('user_management'))


@app.route('/users/reactivate/<int:user_id>', methods=['POST'])
@login_required
def reactivate_user(user_id):
    """Reactivate a user (Admin only)"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied. Only administrators can reactivate users.', 'error')
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(user_id)
    user.is_active = True
    db.session.commit()

    flash(f'User {user.name} has been reactivated', 'success')
    return redirect(url_for('user_management'))


@app.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    """Delete a user permanently (Admin only)"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied. Only administrators can delete users.', 'error')
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(user_id)

    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('user_management'))

    # Check if this is the last admin
    if user.role == UserRole.ADMIN:
        admin_count = User.query.filter_by(role=UserRole.ADMIN).count()
        if admin_count <= 1:
            flash('Cannot delete the last administrator account.', 'error')
            return redirect(url_for('user_management'))

    # Check if user has associated records
    payments_count = Payment.query.filter_by(processed_by=user.id).count()
    assessments_count = FeeAssessment.query.filter_by(assessed_by=user.id).count()
    expenses_count = Expense.query.filter_by(created_by=user.id).count()

    if payments_count > 0 or assessments_count > 0 or expenses_count > 0:
        flash(f'Cannot delete user {user.name} - they have associated records. Consider deactivating instead.', 'error')
        return redirect(url_for('user_management'))

    try:
        user_name = user.name
        db.session.delete(user)
        db.session.commit()

        flash(f'User {user_name} has been permanently deleted', 'success')
        return redirect(url_for('user_management'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'error')
        return redirect(url_for('user_management'))


@app.route('/users/change-role/<int:user_id>', methods=['POST'])
@login_required
def change_user_role(user_id):
    """Quick role change endpoint"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        return jsonify({'error': 'Cannot change your own role'}), 400

    try:
        new_role = request.json.get('role')
        user.role = UserRole(new_role)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Role changed to {new_role}',
            'user': {
                'id': user.id,
                'name': user.name,
                'role': user.role.value
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@app.route('/profile')
@login_required
def user_profile():
    """View current user's profile"""
    # Get user's activity stats
    payments_processed = Payment.query.filter_by(processed_by=current_user.id).count()
    assessments_created = FeeAssessment.query.filter_by(assessed_by=current_user.id).count()
    expenses_created = Expense.query.filter_by(created_by=current_user.id).count()

    stats = {
        'payments_processed': payments_processed,
        'assessments_created': assessments_created,
        'expenses_created': expenses_created
    }

    return render_template('users/profile.html', stats=stats)


@app.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profile():
    """Edit current user's own profile"""
    if request.method == 'POST':
        try:
            current_user.name = request.form['name']

            # Users cannot change their own role or email
            # Email is managed by Google OAuth

            db.session.commit()
            flash('Profile updated successfully', 'success')
            return redirect(url_for('user_profile'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating profile: {str(e)}', 'error')

    return render_template('users/edit_profile.html')

# ===========================
#  MAIN APPLICATION
# ===========================
if __name__ == '__main__':
    with app.app_context():
        # Initialize database and default data
        init_database()

        # Create default admin user if no users exist
        if not User.query.first():
            admin_user = User(
                email='admin@school.com',
                name='System Administrator',
                role=UserRole.ADMIN,
                is_active=True
            )
            db.session.add(admin_user)
            db.session.commit()
            print("Default admin user created: admin@school.com")

    app.run(debug=True)